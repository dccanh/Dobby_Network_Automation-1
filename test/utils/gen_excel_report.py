#!/usr/bin/env python
# -*- coding: utf-8 -*-
import sys
sys.path.append('../')
from selenium import webdriver
import time
import openpyxl
import sys, os
from openpyxl.styles import PatternFill


def export_excel(*argv):
    # parse test result from HTML file
    for file in argv:
        driver = webdriver.Chrome('Driver/chromedriver.exe') # open google chrome
        report_html = os.path.join(os.getcwd(), file)
        driver.get(report_html)

        list_detail_text = driver.find_elements_by_xpath('//a[text()="Detail"]')
        for text_detail in list_detail_text:
            text_detail.click()

        list_popup = driver.find_elements_by_class_name('popup_link')

        # get list of test case
        list_test_case = driver.find_elements_by_class_name('testcase')
        list_detail = driver.find_elements_by_xpath('//td[@colspan = "5"]/a')
        list_real_time = driver.find_elements_by_xpath('//td[@colspan = "5"]//pre')
        # click to display detail information
        for popup in list_popup:
            popup.click()
            time.sleep(1)

        output = []
        num_failed = 0
        # parse data from HTML
        for i in range(len(list_test_case)):
            actual_result = ''
            real_time = list_real_time[i].text.split('Duration: ')[-1][:12]
            output.append([])
            res = str(list_detail[i].text).strip()
            if res == "pass" != -1:
                result = "OK"
                failed_reason = 'As expected'
            elif res == "error":
                failed_reason = '[ERROR] Test scripts were error. Please check it again.'
                result = "ER"
            else:
                result = "NG"
                lines = str(list_real_time[i].text).splitlines()
                for line in lines:
                    if line.find(': [') != -1:
                        failed_reason = line.split(':')[-1].strip()

            output[i] = [list_test_case[i].text, result, failed_reason, actual_result, real_time]

            # write result to excel file
            report_xlsx = 'Report/report.xlsx'
            wb = openpyxl.load_workbook(report_xlsx)

            redFill = PatternFill(start_color='FF0000',
                                  end_color='FF0000',
                                  fill_type='solid')
            GreenFill = PatternFill(start_color='32CD32',
                                  end_color='32CD32',
                                  fill_type='solid')
            OrangeFill = PatternFill(start_color='FF8000',
                                  end_color='FF8000',
                                  fill_type='solid')

            list_sheet_name = wb.sheetnames
            for sheet in list_sheet_name:
                if sheet.lower() in file:
                    ws = wb[sheet]

                    for n in range(2, ws.max_row + 1):
                        for j in range(len(output)):
                            test_case = str(ws.cell(n, 1).value)
                            if output[j][0] == test_case:

                                # actual result
                                ws.cell(n, 10).value = output[j][2]

                                # pass/fail?
                                ws.cell(n, 14).value = output[j][1]

                                # fill actual result text
                                ws.cell(n, 13).value = output[j][3]

                                # real time
                                ws.cell(n, 11).value = output[j][4]

                                if ws.cell(n, 14).value == 'ER':
                                    ws.cell(n, 14).fill = redFill

                                elif ws.cell(n, 14).value == 'NG':
                                    ws.cell(n, 14).fill = OrangeFill
                                    num_failed = num_failed + 1
                                else:
                                    ws.cell(n, 14).fill = GreenFill
                                break

                wb.save(report_xlsx)

        driver.quit()


if __name__ == '__main__':
    for i in range(1, len(sys.argv)):
        export_excel(sys.argv[i])