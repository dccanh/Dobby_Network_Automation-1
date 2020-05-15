#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os, sys
sys.path.append('../../')
import sys
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.remote.webelement import WebElement

import random

import json

import gspread

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
# from datetime import datetime
# from datetime import date
# from faker import Faker
# from Helper.t10x.config.read_config import *
from Helper.t10x.config.elements import *
from Helper.t10x.config.captcha import *
# import base64
from oauth2client.service_account import ServiceAccountCredentials
# from Helper.t10x.ls_path import *
from Helper.t10x.secure_crt.common import *
import re
# import xml.etree.ElementTree as ET
from winreg import *
def save_config(config_path, section, option, value):
    config = configparser.RawConfigParser()
    config.read(config_path)
    if not config.has_section(str(section).upper()):
        config.add_section(str(section).upper())
    config.set(str(section).upper(), str(option), str(value))
    with open(config_path, 'w', encoding='utf-8') as config_file:
        config.write(config_file)



def get_config(section, option, config_dir=config_path):
    if not os.path.exists(config_dir):
        print("The config file not exist. Exit!!!")
        return

    config = configparser.RawConfigParser()
    config.read(config_dir)

    if config.has_option(str(section).upper(), option):
        return config.get(str(section).upper(), option)
    else:
        return

serial_num = get_config('GENERAL', 'serial_number')
save_config(config_path, 'ACCOUNT', 'default_pw', serial_num)


def next_available_row(sheet):
    str_list = list(filter(None, sheet.col_values(1)))  # fastest
    return str(len(str_list) + 1)


def write_to_excel(key, list_steps, func_name, duration, time_stamp=0):
    import openpyxl
    ls = subprocess.check_output('tasklist')
    if b'EXCEL.EXE' in ls:
        os.system("taskkill /f /im EXCEL.EXE")

    excel_file = report_offline_path
    #
    # from datetime import datetime
    # unix_string = datetime.today().strftime('%Y-%m-%d_%H_%M_%S')
    # customize_report_path = os.path.join(get_config('REPORT', 'report_path'), f"Report_automation_{unix_string}.xlsx")
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    for i in range(2, ws.max_row + 2):
        if ws.cell(i, 1).value is None:
            ws.cell(row=i, column=1).value = key
            ws.cell(row=i, column=2).value = func_name
            # Fill result
            if '[Fail]' in str(list_steps):
                ws.cell(row=i, column=3).value = 'FAIL'
            else:
                if '[END TC]' not in list_steps:
                    ws.cell(row=i, column=3).value = 'FAIL'
                    list_steps.append('Can not execute next step ...')
                else:
                    ws.cell(row=i, column=3).value = 'PASS'

            # Fill duration
            ws.cell(row=i, column=4).value = duration
            # Fill step
            steps = ''
            for j in list_steps:
                steps = steps + (str(j) + '\n')
            ws.cell(row=i, column=5).value = steps

            ws.cell(row=i, column=6).value = time_stamp
            ws.cell(row=1, column=5).value = get_config('REPORT', 'sheet_name')
            # Save file
            wb.save(excel_file)
            # wb.save(customize_report_path)
            break

def write_to_excel_tmp(key, list_steps, func_name):
    import openpyxl
    ls = subprocess.check_output('tasklist')
    if b'EXCEL.EXE' in ls:
        os.system("taskkill /f /im EXCEL.EXE")

    excel_file = report_offline_path
    wb = openpyxl.load_workbook(excel_file)
    # wb.active = 2
    ws = wb.active

    for i in range(2, ws.max_row + 2):
        if ws.cell(i, 1).value is None:
            ws.cell(row=i, column=1).value = key
            ws.cell(row=i, column=2).value = func_name
            # Fill result
            if '[Fail]' in str(list_steps):
                ws.cell(row=i, column=3).value = 'FAIL'
            else:
                if '[END TC]' not in list_steps:
                    ws.cell(row=i, column=3).value = 'FAIL'
                    list_steps.append('Can not execute next step ...')
                else:
                    ws.cell(row=i, column=3).value = 'PASS'
            # Fill step
            steps = '\n'.join(list_steps)
            # for j in list_steps:
            #     steps = steps + (str(j) + '\n')
            ws.cell(row=i, column=5).value = steps

            # Save file
            wb.save(excel_file)
            break


# def write_ggsheet(key, list_steps, func_name, duration, time_stamp=0):
#     scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
#     creds = ServiceAccountCredentials.from_json_keyfile_name(gg_credential_path, scope)
#     client = gspread.authorize(creds)
#     # sheet = client.open("[DOB] Report Automation").get_worksheet(1)
#     get_gg_sheet_name = get_config('REPORT', 'sheet_name')
#     sheet = client.open("[DOB] Report Automation").worksheet(get_gg_sheet_name)
#     next_row = next_available_row(sheet)
#     sheet.update_acell("A{}".format(next_row), key)
#     sheet.update_acell("B{}".format(next_row), func_name)
#     list_steps = '\n'.join(list_steps)
#     if '[Fail]' in str(list_steps):
#         sheet.update_acell("C{}".format(next_row), 'FAIL')
#     else:
#         # if '[END TC]' not in str(list_steps):
#         if '[END TC]' not in list_steps:
#             list_steps += ('\nCan not execute next step ...')
#             sheet.update_acell("C{}".format(next_row), 'FAIL')
#         else:
#             sheet.update_acell("C{}".format(next_row), 'PASS')
#     sheet.update_acell("D{}".format(next_row), duration)
#     sheet.update_acell("E{}".format(next_row), str(list_steps))
#     sheet.update_acell("H{}".format(next_row), str(time_stamp))


def write_ggsheet(key, list_steps, func_name, duration, time_stamp=0):
    from gspread.models import Cell
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    creds = ServiceAccountCredentials.from_json_keyfile_name(gg_credential_path, scope)
    client = gspread.authorize(creds)
    get_gg_sheet_name = get_config('REPORT', 'sheet_name')
    sheet = client.open("[DOB] Report Automation").worksheet(get_gg_sheet_name)
    next_row = next_available_row(sheet)
    list_steps = '\n'.join(list_steps)
    cells = list()
    cells.append(Cell(int(next_row), 1, key))
    cells.append(Cell(int(next_row), 2, func_name))
    if '[Fail]' in str(list_steps):
        cells.append(Cell(int(next_row), 3, 'FAIL'))
    else:

        if '[END TC]' not in list_steps:
            list_steps += ('\nCan not execute next step ...')
            cells.append(Cell(int(next_row), 3, 'FAIL'))
        else:

            cells.append(Cell(int(next_row), 3, 'PASS'))
    cells.append(Cell(int(next_row), 4, duration))
    cells.append(Cell(int(next_row), 5, "\n".join(list_steps)))
    cells.append(Cell(int(next_row), 8, str(time_stamp)))
    sheet.update_cells(cells)

def assert_list(list_actual_result, list_expected_result):
    actual_fail = []
    expected_fail = []
    for a, e in zip(list_actual_result, list_expected_result):
        if str(a) != str(e):
            actual_fail.append(str(a))
            expected_fail.append(str(e))
    if len(actual_fail) == 0 and len(expected_fail) == 0 and len(list_actual_result) == len(list_expected_result):
        return {"result": True, "actual": [], "expected": []}
    return {"result": False, "actual": list_actual_result, "expected": list_expected_result}


def get_func_name():
    import inspect
    return inspect.stack()[1][3]


def convert_type_number(num):
    if '.0' in str(num):
        return str(num).split('.0')[0]
    return str(num)


# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~REQUEST GENERAL  BLOCK~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
def wait_visible(driver, css_element):
    visible = driver.find_elements_by_css_selector(css_element)
    count = 0
    while len(visible) == 0:
        time.sleep(1)
        visible = driver.find_elements_by_css_selector(css_element)
        count += 1
        if count == 300:
            driver.quit()
            os.system("[Fail] Loaded time out.")
            return False
    else:
        return True


def wait_popup_disappear(driver, pop_up_element):
   visible = driver.find_elements_by_css_selector(pop_up_element)
   count =0
   while len(visible) != 0:
       time.sleep(1)
       visible = driver.find_elements_by_css_selector(pop_up_element)
       count += 1
       if count == 300:
           driver.quit()
           os.system("[Fail] Loaded time out.")
           return False
   else:
       return True


def random_list_drop_down_with_condition(driver, css_option_condition):
    project_options = driver.find_elements_by_css_selector(css_option_condition)
    choice = random.choice(project_options)
    ActionChains(driver).move_to_element(choice).perform()
    option_value = choice.text
    choice.click()
    return {"chosenValue": option_value}


def random_list_drop_down_regis_emp(driver, css_option_condition):
    project_options = driver.find_elements_by_css_selector(css_option_condition)
    if len(project_options):
        choice = random.choice(project_options)
        ActionChains(driver).move_to_element(choice).perform()
        option_value = choice.text
        choice.click()
        return {"chosenValue": option_value}
    else:
        return False


def read_column(driver, col):
    # col = name
    elem1 = driver.find_element_by_xpath("//span[contains(text(),'" + col + "')]/../../..")
    lst = driver.find_elements_by_xpath("//div[@class='ant-table-scroll']/div[1]/table/thead/tr/th")
    a = lst.index(elem1) + 1

    time.sleep(2)
    list_pages = driver.find_elements_by_xpath('//ul//li//a')
    data = []
    info_each_row = []
    if len(list_pages) != 0:
        num_pages = int(list_pages[-2].text)
        for num in range(1, num_pages + 1):
            num_rows = driver.find_elements_by_css_selector('.ant-table-scroll tr.ant-table-row-level-0')
            for row in range(0, len(num_rows)):
                if a > 2:
                    num_cells = driver.find_elements_by_css_selector(
                        "div.ant-table-scroll> * tr>td:nth-child(" + str(a) + ")")
                    time.sleep(0.01)
                    data.append(num_cells[row].text)
                else:
                    num_cells = driver.find_elements_by_css_selector(
                        ".ant-table-fixed-left tr>td:nth-child(" + str(a) + ")")
                    time.sleep(0.01)
                    data.append(num_cells[row].text)
            list_pages[-1].click()
            time.sleep(2)
    else:
        total_result = int(driver.find_element_by_css_selector('h3.text-primary.mb-0.mr-2').text.split(': ')[1])
        if total_result != 0:
            num_rows = driver.find_elements_by_css_selector('.ant-table-scroll tr.ant-table-row-level-0')
            for row in range(0, len(num_rows)):
                if a > 2:
                    num_cells = driver.find_elements_by_css_selector(
                        ".ant-table-scroll tr>td:nth-child(" + str(a) + ")")
                    ActionChains(driver).move_to_element(num_cells[row]).perform()
                    data.append(num_cells[row].text)
                else:
                    num_cells = driver.find_elements_by_css_selector(
                        ".ant-table-fixed-left tr>td:nth-child(" + str(a) + ")")
                    ActionChains(driver).move_to_element(num_cells[row]).perform()
                    data.append(num_cells[row].text)
        else:
            data = 'No data'
    return data


def scroll_to(driver, element):
    coordinates = element.location_once_scrolled_into_view  # returns dict of X, Y coordinates
    driver.execute_script('window.scrollTo({}, {});'.format(coordinates['x'], coordinates['y']))


def get_result_command_from_server(url_ip, filename='1'):
    driver2 = webdriver.Chrome(driver_path)
    url_result = url_ip + '/' + filename
    time.sleep(1)
    driver2.get(url_result)
    time.sleep(2)
    get_all_text = driver2.find_element_by_css_selector('html>body>pre').text
    command_result = '{' + get_all_text.split('@@@ PRINT OBJLIST START @@@')[1].split(' @@@@@@ PRINT OBJLIST END @@@@@@')[0]
    fit_result = command_result.split('''   },\n  },\n }''')[0]+'}}}'

    info = json.loads(fit_result)
    username = info['Device.Users.User.2']['Username']['paramValue']
    password = info['Device.Users.User.2']['Password']['paramValue']

    save_config(config_path, 'ACCOUNT', 'user', username)
    save_config(config_path, 'ACCOUNT', 'password', password)
    driver2.quit()
    return {'userName': username, 'passWord': password}

def get_result_command_from_server_api(url_ip, _name='1'):
    url = url_ip + '/' + _name
    print(url)
    _res = requests.get(url)
    # if _res.status_code != 200:
    #     url_login = get_config('URL', 'url')
    #     filename_2 = 'account2.txt'
    #     command_2 = 'capitest get Device.Users.User.2. leaf'
    #     run_cmd(command_2, filename_2)
    #     time.sleep(5)
    #     url = url_login + '/' + filename_2
    #     _res = requests.get(url)
    #     if _res.status_code != 200:
    #         url_login = get_config('URL', 'url')
    #         filename_1 = '1'
    #         command = 'factorycfg.sh -a'
    #         run_cmd(command, filename_1)
    #         # Wait 5 mins for factory
    #         time.sleep(150)
    #         wait_DUT_activated(url_login)
    #         wait_ping('192.168.1.1')
    #         time.sleep(3)
    #         filename_2 = 'account2.txt'
    #         command_2 = 'capitest get Device.Users.User.2. leaf'
    #         run_cmd(command_2, filename_2)
    #         time.sleep(5)
    #         url = url_login + '/' + filename_2
    #         _res = requests.get(url)

    get_all_text = _res.text
    print("Status Code: " + str(_res.status_code))
    print("Your Serial Number is: " + get_config('GENERAL', 'serial_number'))
    command_result = '{' + get_all_text.split('@@@ PRINT OBJLIST START @@@')[1].split(' @@@@@@ PRINT OBJLIST END @@@@@@')[0]
    fit_result = command_result.split('},\n\t },\n }')[0]+'}}}'

    info = json.loads(fit_result)
    username = info['Device.Users.User.2']['Username']['paramValue']
    password = info['Device.Users.User.2']['Password']['paramValue']
    print(username, password)
    save_config(config_path, 'ACCOUNT', 'user', username)
    save_config(config_path, 'ACCOUNT', 'password', password)
    return {'userName': username, 'passWord': password}


def login(driver, url_login='', user_request='', pass_word=''):
    if url_login == '':
        url_login = get_config('URL', 'url')

    if user_request == '':
        user_request = get_config('ACCOUNT', 'user')

    if pass_word == '':
        pass_word = get_config('ACCOUNT', 'password')

    print(url_login)
    print(user_request)
    print(pass_word)

    # a = call_api_login(user_request, pass_word, url=url_login)
    # print(a)


    # user_request = get_config('ACCOUNT', 'user')
    # pass_word = get_config('ACCOUNT', 'password')
    time.sleep(1)
    driver.get(url_login)
    time.sleep(2)
    if driver.current_url == 'http://ww38.dearmyextender.net/':
        save_config(config_path, 'URL', 'url', 'http://dearmyrouter.net')
        url_login = get_config('URL', 'url')
        driver.get(url_login)

    driver.find_elements_by_css_selector(lg_user)[-1].send_keys(user_request)
    time.sleep(1)
    driver.find_elements_by_css_selector(lg_password)[-1].send_keys(pass_word)
    time.sleep(1)
    # Captcha
    captcha_src = driver.find_element_by_css_selector(lg_captcha_src).get_attribute('src')
    captcha_text = get_captcha_string(captcha_src)
    driver.find_element_by_css_selector(lg_captcha_box).send_keys(captcha_text)
    time.sleep(1)
    driver.find_elements_by_css_selector(lg_btn_login)[-1].click()

    # If login Fail. Get USER and PW again
    msg_error = driver.find_element_by_css_selector(lg_msg_error).text
    if msg_error != '':
        filename_2 = 'account.txt'
        command_2 = 'capitest get Device.Users.User.2. leaf'
        run_cmd(command_2, filename_2)
        time.sleep(3)
        get_result_command_from_server(url_ip=url_login, filename=filename_2)

        user_request = get_config('ACCOUNT', 'user')
        pass_word = get_config('ACCOUNT', 'default_pw')

        time.sleep(1)
        driver.get(url_login)
        time.sleep(2)
        driver.find_elements_by_css_selector(lg_user)[-1].send_keys(user_request)
        time.sleep(1)
        driver.find_elements_by_css_selector(lg_password)[-1].send_keys(pass_word)
        time.sleep(1)
        # Captcha
        captcha_src = driver.find_element_by_css_selector(lg_captcha_src).get_attribute('src')
        captcha_text = get_captcha_string(captcha_src)
        driver.find_element_by_css_selector(lg_captcha_box).send_keys(captcha_text)
        time.sleep(1)
        driver.find_elements_by_css_selector(lg_btn_login)[-1].click()
        time.sleep(5)
    wait_popup_disappear(driver, dialog_loading)
    time.sleep(1)
    # Check Privacy Policy
    time.sleep(2)
    policy_popup = driver.find_elements_by_css_selector(lg_privacy_policy_pop)
    if len(policy_popup):
        wait_popup_disappear(driver, icon_loading)
        wait_popup_disappear(driver, icon_loading)
        time.sleep(3)
        ActionChains(driver).move_to_element(policy_popup[0]).click().send_keys(Keys.ARROW_DOWN).perform()
        time.sleep(1)
        driver.find_element_by_css_selector(btn_ok).click()
        time.sleep(3)


def grand_login(driver, url_login='', user_request='', pass_word=''):
    login(driver, url_login, user_request, pass_word)
    # wait_popup_disappear(driver, dialog_loading)
    time.sleep(1)
    # Goto Homepage
    if len(driver.find_elements_by_css_selector(lg_welcome_header)) != 0:
        handle_winzard_welcome(driver)
        wait_popup_disappear(driver, dialog_loading)
    check_ota_auto_update(driver)
    time.sleep(1)


def get_url_ipconfig(ipconfig_field='Default Gateway'):
    cmd = 'ipconfig'
    write_cmd = subprocess.check_output(cmd, encoding='oem')
    split_result = [i.strip() for i in write_cmd.splitlines()]
    default_gw = [i for i in split_result if i.startswith(ipconfig_field)]
    ip = [i.split(':')[1].strip() for i in default_gw if i.split(':')[1].strip()][0]
    # url_ = 'http://'+ip
    # save_config(config_path, 'URL', 'url', url_)
    return ip


def get_value_from_ipconfig(block, field):
    import subprocess
    cmd = 'ipconfig/all'
    write_cmd = subprocess.check_output(cmd, shell=True)
    write_cmd = write_cmd.decode('utf8')
    # print(write_cmd)
    if block in write_cmd:
        write_cmd = write_cmd.split(block)[1].split('\r\n\r\n')[1]
        split_result = [i.strip() for i in write_cmd.splitlines()]
        field_row = [i for i in split_result if i.startswith(field)]
        if len(field_row):
            field_value = [i.split(':')[1].strip() for i in field_row if i.split(':')[1].strip()][0]
        else:
            field_value = 'Block or field error.'
        return field_value
    else:
        return 'Block or field error.'


def goto_menu(driver, parent_tab, child_tab):
    ActionChains(driver).move_to_element(driver.find_element_by_css_selector(parent_tab)).click().perform()
    time.sleep(1)
    wait_popup_disappear(driver, dialog_loading)
    if child_tab != 0:
        driver.find_element_by_css_selector(child_tab).click()
        time.sleep(0.5)
        wait_popup_disappear(driver, dialog_loading)


def detect_current_menu(driver):
    # Detect current active tab in T10x
    main_tab = driver.find_element_by_css_selector('#parent-menu>a.active.open')
    ActionChains(driver).move_to_element(main_tab).perform()
    main_tab = main_tab.text

    child_tab = driver.find_elements_by_css_selector('#sub-menu>.active')
    if not len(child_tab):
        child_tab = 0
    else:
        child_tab = child_tab[0].text
    return main_tab, child_tab


def base64encode(user, pw):
    import hashlib, binascii, base64
    salt = base64.b64encode(('hmx#cpe@pbkdf2*SALT!' + user).encode('utf-8'))

    # salt = binascii.hexlify(hash0_pw)
    if salt.decode('utf8') != '':
        mode = 'sha512'
        iterations = 1000
        dklen = 64

        hash0_pw = hashlib.pbkdf2_hmac(mode, pw.encode('utf-8'), salt, iterations, dklen)
        # hash1_pw = binascii.hexlify(hash0_pw)
        hash1_pw = base64.b64encode(hash0_pw)

        hash2_pw = hash1_pw.decode('utf-8')
        initial_data = 'HS:' + user + ':' + hash2_pw
        byte_string = initial_data.encode('utf-8')
        encoded_data = base64.b64encode(byte_string)
        return encoded_data.decode('ascii')
    else:
        initial_data = 'HS:' + user + ':' + pw
        byte_string = initial_data.encode('utf-8')
        encoded_data = base64.b64encode(byte_string)

        return encoded_data.decode('ascii')

def base64encodev2(user, pw):
    import hashlib, binascii, base64

    initial_data = 'HS:' + user + ':' + pw
    byte_string = initial_data.encode('utf-8')
    encoded_data = base64.b64encode(byte_string)

    return encoded_data.decode('ascii')

def send_request(url, method, headers, body, timeout=120):
    result = requests.models.Response()
    try:
        if method.upper() == 'GET':
            result = requests.get(url=url, headers=headers, json=body, timeout=timeout)
        elif method.upper() == 'POST':
            result = requests.post(url=url, headers=headers, json=body, timeout=timeout)
        elif method.upper() == 'PUT':
            result = requests.put(url=url, headers=headers, json=body, timeout=timeout)
        elif method.upper() == 'DELETE':
            result = requests.put(url=url, headers=headers, json=body, timeout=timeout)
    except Exception:
        result.status_code = 404
    return result


def call_api_login(user, pw, url=''):
    """Method: POST; Require: user, pw; Return: JSON data"""
    if url == '':
        url = get_config('URL', 'url')
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pw)
    }
    print(base64encode(user, pw))
    res = requests.post(url=url_login, json=data)
    print(res)
    if res.status_code != 200:

        user = get_config('ACCOUNT', 'user')
        pw = get_config('ACCOUNT', 'default_pw')
        data = {
            "userName": user,
            "password": base64encode(user, pw)
        }
        res = requests.post(url=url_login, json=data)
        if res.status_code == 200:
            save_config(config_path, 'ACCOUNT', 'password', pw)

    json_data = json.loads(res.text)
    return json_data


def get_token(user, pw):
    token = call_api_login(user, pw)["accessToken"]
    return token


def call_api(url, method, body, token):

    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res_data = send_request(url, method, headers, body)
    if res_data.status_code != 404:
        data = json.loads(res_data.text)
    else:
        data = {"body": "404", "statusCode": 404}
    return data


def check_page_exist(url):
    # api_url = ipv4 + "/index.html"
    try:
        res = requests.get(url=url, timeout=1)
        if (res.status_code == 200):
            return True
    except Exception:
        return False


def wait_DUT_activated(url):
    count = 0
    url = url + "/index.html"
    while True:
        count += 1

        print(url)
        if check_page_exist(url):
            return True
        else:
            print("          WAITING FOR THE DEVICE ACTIVED... | RE-TRY TIMES: " + str(count))
            time.sleep(1)
            if (count % 180 == 0):
                return False


def ping_to_url(url):
    TIME_OUT = 'time out'
    result = subprocess.check_output('ping ' + url)
    if TIME_OUT not in result.decode('utf8'):
        # Check there is no time out in pinged result > Successfully
        return True
    return False


def wait_ping(url):
    TIME_OUT = 'Request time out'
    count = 0
    while True:
        result = subprocess.check_output('ping ' + url, shell=True)
        time.sleep(1)
        count += 1
        if TIME_OUT not in result.decode('utf8'):
            break
        else:
            if count == 300:
                print('Wait more than 300s')
                break


def scan_wifi(wifi_name=None):
    cmd = 'netsh wlan show networks'
    write_cmd = subprocess.check_output(cmd, encoding='oem')
    ls_header = [i for i in write_cmd.splitlines() if i.startswith('SSID')]
    ls_wifi_name = [i.split(': ')[1] for i in ls_header]
    if wifi_name is not None:
        while True:
            write_cmd = subprocess.check_output(cmd, encoding='oem')
            ls_header = [i for i in write_cmd.splitlines() if i.startswith('SSID')]
            ls_wifi_name = [i.split(': ')[1] for i in ls_header]
            if wifi_name in ls_wifi_name:
                return True
            else:
                time.sleep(1)

    return ls_wifi_name


def change_nw_profile(wifi_xml_path, field, value):
    """
    :param wifi_xml_path: wifi_2g_path or wifi_5g_path
    :param field: Fields name: SSID, Security or Password
    :param value: Value of field
    :return: overwrite wifi_xml_path new value
    """
    value = value.replace('&', '&amp;')
    value = value.replace('>', '&gt;')
    value = value.replace('<', '&lt;')
    trans_dict = {"Ssid": "name",
                  "Security": "authentication",
                  "Encryption": "encryption",
                  "Password": "keyMaterial"}
    temp_path = config_dir + '\\temp.txt'

    os.rename(wifi_xml_path, temp_path)
    time.sleep(1)
    with open(temp_path) as f:
        data = f.read()
        rows = [i.strip() for i in data.splitlines() if i.strip().startswith('<' + trans_dict[field] + '>')]
        # old = a[0].split('<' + trans_dict[field] + '>')[1].split('</' + trans_dict[field] + '>')[0]
        for r in rows:
            old = r.split('<' + trans_dict[field] + '>')[1].split('</' + trans_dict[field] + '>')[0]
            data = data.replace(old, value)
        # b = data.replace(old, value)
        f.close()
    with open(temp_path, 'w') as f:
        f.write(data)
        f.close()
    os.rename(temp_path, wifi_xml_path)


def connect_wifi_from_xml(wifi_xml_path):
    temp_path = config_dir + '\\temp.txt'
    os.rename(wifi_xml_path, temp_path)
    with open(temp_path) as f:
        data = f.read()
        name = [i.strip() for i in data.splitlines() if i.strip().startswith('<name>')]
        _name = name[0].split('<name>')[1].split('</name>')[0]
        f.close()
    os.rename(temp_path, wifi_xml_path)

    # Add network profile
    cmd_add_profile = 'netsh wlan add profile filename="'+wifi_xml_path+'"'
    print(cmd_add_profile)
    os.system(cmd_add_profile)
    time.sleep(2)
    # Connect that name
    cmd_connect = 'netsh wlan connect ssid="'+_name+'" name="'+_name+'"'
    print(cmd_connect)
    os.system(cmd_connect)


def setting_wireless_security(block_left_right, secur_type=None, encryption=None, key_type=None):
    # block_ = driver.find_element_by_css_selector(block_element)

    security_ = block_left_right.find_element_by_css_selector(secure_value_field)
    security_.click()
    ls_security_ = security_.find_elements_by_css_selector(secure_value_in_drop_down)
    time.sleep(0.5)
    for o in ls_security_:
        if o.get_attribute('option-value') == secur_type:
            o.click()
            break

    if encryption is not None:
        encryption_ = block_left_right.find_element_by_css_selector(encryption_value_field)
        encryption_.click()
        ls_encryption_ = encryption_.find_elements_by_css_selector(secure_value_in_drop_down)
        time.sleep(0.5)
        for o in ls_encryption_:
            if o.get_attribute('option-value') == encryption:
                o.click()
                break

    if key_type is not None:
        key_type_ = block_left_right.find_element_by_css_selector(encryption_value_field)
        key_type_.click()
        ls_key_type_ = key_type_.find_elements_by_css_selector(secure_value_in_drop_down)
        time.sleep(0.5)
        for o in ls_key_type_:
            if o.get_attribute('option-value') == key_type:
                o.click()
                break


def apply_process(driver, block_left_right):
    time.sleep(0.2)
    block_left_right.find_element_by_css_selector(apply).click()
    time.sleep(1)
    wait_popup_disappear(driver, dialog_loading)
    driver.find_element_by_css_selector(btn_ok).click()
    wait_popup_disappear(driver, dialog_loading)
    time.sleep(2)


def checkMACAddress(x):
    if re.match("[0-9a-f]{2}([-:])[0-9a-f]{2}(\\1[0-9a-f]{2}){4}$", x.lower()):
        return True
    return False


def checkIPAddress(x):
    if re.match("^(([0-9]|[1-9][0-9]|1[0-9]{2}|2[0-4][0-9]|25[0-5])\.){3}([0-9]|[1-9][0-9]|1[0-9]{2}|2[0-4][0-9]|25[0-5])$", x):
        return True
    return False


def handle_winzard_welcome(driver, NEW_PASSWORD='abc123', exp_language='English'):
    exp_time_zone = '(GMT+07:00) Bangkok, Ho Chi Minh, Phnom Penh, Vientiane'
    if driver.find_element_by_css_selector('.language-dialog-select input').get_attribute('value') != exp_language:
        # Click to Language
        driver.find_element_by_css_selector(welcome_language).click()
        time.sleep(3)

        # Choose Language
        ls_time_zone = driver.find_elements_by_css_selector(welcome_list_language)
        for t in ls_time_zone:
            ActionChains(driver).move_to_element(t).perform()
            if t.text == exp_language:
                t.click()
                break
        time.sleep(1)

    if driver.find_element_by_css_selector('.datetime-dialog-select input').get_attribute('value') != exp_time_zone:
        # Click to time zone
        driver.find_element_by_css_selector(welcome_time_zone).click()
        time.sleep(1)

        # Choose time zone in drop down: Vn zone GMT +7
        ls_time_zone = driver.find_elements_by_css_selector(welcome_list_time_zone)
        for t in ls_time_zone:
            ActionChains(driver).move_to_element(t).perform()
            if t.text == exp_time_zone:
                t.click()
                break

    time.sleep(1)
    # Click start btn
    driver.find_element_by_css_selector(welcome_start_btn).click()
    time.sleep(3)
    wait_visible(driver, welcome_change_pw_fields)
    change_pw_fields = driver.find_elements_by_css_selector(welcome_change_pw_fields)

    # A list contain values: Current Password, New Password, Retype new pw
    ls_change_pw_value = [get_config('ACCOUNT', 'password'), NEW_PASSWORD, NEW_PASSWORD]
    for p, v in zip(change_pw_fields, ls_change_pw_value):
        ActionChains(driver).move_to_element(p).click().send_keys(v).perform()
        time.sleep(0.5)

    # Next Change pw
    time.sleep(1)
    wait_visible(driver, welcome_next_btn)
    next_btn = driver.find_element_by_css_selector(welcome_next_btn)
    if not next_btn.get_property('disabled'):
        next_btn.click()
    time.sleep(3)

    while True:
        time.sleep(1)
        wait_visible(driver, welcome_next_btn)
        next_btn = driver.find_element_by_css_selector(welcome_next_btn)
        if not next_btn.get_property('disabled'):
            next_btn.click()
        time.sleep(3)

        if len(driver.find_elements_by_css_selector(welcome_let_go_btn)) > 0:
            break

    time.sleep(3)
    driver.find_element_by_css_selector(welcome_let_go_btn).click()
    # Write config
    save_config(config_path, 'ACCOUNT', 'password', NEW_PASSWORD)
    wait_popup_disappear(driver, dialog_loading)
    time.sleep(1)
    wait_visible(driver, home_view_wrap)



def ping_to_address(PING_ADDRESS, PING_TIMES=4):
    import subprocess
    result = subprocess.check_output(f'ping {str(PING_ADDRESS)} -n {str(PING_TIMES)}', shell=True)

    for r in result.splitlines():
        if r.decode('utf8').strip().startswith('Packets:'):
            result_row = r.decode('utf8').strip()
            packets_loss_rate = result_row.split('(')[1].split('%')[0]
            break
    return {"packet_loss_rate": float(packets_loss_rate)}


def network_interface_action(interface='Ethernet', action='enable'):
    # Interface option Wi-Fi / Ethernet
    # Action: enabled/disabled
    import ctypes

    print("interface: " + interface + " | action: " + action)

    cmd = f'netsh interface set interface name="{interface}" admin={action}'

    def is_admin():
        try:
            return ctypes.windll.shell32.IsUserAnAdmin()
        except:
            return False

    if is_admin():
        # Code of your program here
        subprocess.call(cmd, shell=True)
    else:
        # Re-run the program with admin rights
        parameters = ""
        for i in range(1, len(sys.argv)):
            parameters = parameters + " \"" + str(sys.argv[i]) + "\""
        ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, __file__ + parameters, None, 1)


def api_change_wifi_setting(url_change, new_wifi_name='', get_only_mac=False, ):
    user = get_config('ACCOUNT', 'user')
    pw = get_config('ACCOUNT', 'password')
    token = get_token(user, pw)
    wifi_default_pw = 'humax_' + get_config('GENERAL', 'serial_number')
    hash_pw = base64encodev2(user, wifi_default_pw)
    get_mac = call_api(url_change, 'GET', body='', token=token)['macAddress']
    if get_only_mac:
        return get_mac
    else:
        if new_wifi_name == '':
            new_wifi_name = '_'.join(['wifi', get_mac.replace(':', '_')])
        data = {
            "active": True,
            "name": new_wifi_name,
            "macAddress": get_mac,
            "hiddenSSID": False,
            "APIsolate": False,
            "webUIAccess": True,
            "internetOnly": False,
            "accessControl": False,
            "security": {
                "type": "WPA2/WPA-PSK",
                "personal": {
                    "password": hash_pw,
                    "encryption": "AES/TKIP",
                    "groupKey": 3600
                },
                "enterprise": None,
                "wep": None
            },
            "index": 0
        }
        call_api(url_change, 'PUT', body=data, token=token)
        return new_wifi_name


def write_data_to_xml(xml_path, new_name='',
                      new_pw='humax_'+get_config('GENERAL', 'serial_number'),
                      new_secure='',
                      new_encryption='',
                      new_key_type=''):
    with open(xml_path) as wf:
        data = wf.read()
        if new_name != '':
            rows_name = [i.strip() for i in data.splitlines() if i.strip().startswith('<name>')]
            for r in rows_name:
                old = r.split('<name>')[1].split('</name>')[0]
                data = data.replace(old, new_name)

            rows_hex = [i.strip() for i in data.splitlines() if i.strip().startswith('<hex>')][0]
            old = rows_hex.split('<hex>')[1].split('</hex>')[0]
            data = data.replace(old, new_name.encode('utf-8').hex().upper())

        if new_pw != '':
            rows_hex = [i.strip() for i in data.splitlines() if i.strip().startswith('<keyMaterial>')][0]
            old = rows_hex.split('<keyMaterial>')[1].split('</keyMaterial>')[0]
            data = data.replace(old, new_pw)

        if new_secure != '':
            rows_hex = [i.strip() for i in data.splitlines() if i.strip().startswith('<authentication>')][0]
            old = rows_hex.split('<authentication>')[1].split('</authentication>')[0]
            data = data.replace(old, new_secure)

        if new_encryption != '':
            rows_hex = [i.strip() for i in data.splitlines() if i.strip().startswith('<encryption>')][0]
            old = rows_hex.split('<encryption>')[1].split('</encryption>')[0]
            data = data.replace(old, new_encryption)

        if new_key_type != '':
            rows_hex = [i.strip() for i in data.splitlines() if i.strip().startswith('<keyType>')][0]
            old = rows_hex.split('<keyType>')[1].split('</keyType>')[0]
            data = data.replace(old, new_key_type)

        wf.close()

    with open(xml_path, 'w') as f:
        f.write(data)
        f.close()


def write_data_to_none_secure_xml(xml_path, new_name='',
                      new_secure='',
                      new_encryption=''):
    with open(xml_path) as wf:
        data = wf.read()
        if new_name != '':
            rows_name = [i.strip() for i in data.splitlines() if i.strip().startswith('<name>')]
            for r in rows_name:
                old = r.split('<name>')[1].split('</name>')[0]
                data = data.replace(old, new_name)

            rows_hex = [i.strip() for i in data.splitlines() if i.strip().startswith('<hex>')][0]
            old = rows_hex.split('<hex>')[1].split('</hex>')[0]
            data = data.replace(old, new_name.encode('utf-8').hex())

        if new_secure != '':
            rows_hex = [i.strip() for i in data.splitlines() if i.strip().startswith('<authentication>')][0]
            old = rows_hex.split('<authentication>')[1].split('</authentication>')[0]
            data = data.replace(old, new_secure)

        if new_encryption != '':
            rows_hex = [i.strip() for i in data.splitlines() if i.strip().startswith('<encryption>')][0]
            old = rows_hex.split('<encryption>')[1].split('</encryption>')[0]
            data = data.replace(old, new_encryption)
        wf.close()

    with open(xml_path, 'w') as f:
        f.write(data)
        f.close()


def download_destination_path():
    with OpenKey(HKEY_CURRENT_USER, 'SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders') as key:
        Downloads = QueryValueEx(key, '{374DE290-123F-4565-9164-39C4925E467B}')[0]

    return Downloads


def change_pw(driver, pw):
    current_pw = get_config('ACCOUNT', 'password')
    ls_pw_box = driver.find_elements_by_css_selector(' '.join([dialog_content, password_input_cls]))
    # Current pw
    ActionChains(driver).move_to_element(ls_pw_box[0]).click().send_keys(current_pw).perform()
    time.sleep(0.2)
    # New pw
    ActionChains(driver).move_to_element(ls_pw_box[1]).click().send_keys(pw).perform()
    time.sleep(0.2)
    # Retype new pw
    ActionChains(driver).move_to_element(ls_pw_box[2]).click().send_keys(pw).perform()
    time.sleep(0.2)
    # CLick to other place
    driver.find_element_by_css_selector(apply).click()
    time.sleep(0.2)
    wait_popup_disappear(driver, dialog_loading)
    time.sleep(1)


def check_ota_auto_update(driver):
    if len(driver.find_elements_by_css_selector(ele_upgrade_server_popup)) > 0:
        driver.find_element_by_css_selector(btn_cancel).click()
        time.sleep(1)
        # turn Off OTA
        # Actions Systems > Backup
        system_button = driver.find_element_by_css_selector(system_btn)
        ActionChains(driver).move_to_element(system_button).click().perform()
        time.sleep(0.5)
        driver.find_element_by_css_selector(ele_sys_firmware_update).click()
        time.sleep(0.5)
        # Click disable auto update
        driver.find_element_by_css_selector('.dialog-content .toggle-button').click()
        wait_popup_disappear(driver, dialog_loading)
        driver.find_element_by_css_selector(ele_close_button).click()
        time.sleep(1)


def wireless_get_default_ssid(driver, label_name):
    edit_2g_label = driver.find_elements_by_css_selector(label_name_in_2g)
    edit_2g_fields = driver.find_elements_by_css_selector(wrap_input)
    for l, f in zip(edit_2g_label, edit_2g_fields):
        # Connection type
        if l.text == label_name:
            default_ssid_2g_value = f.find_element_by_css_selector(input).get_attribute('value')
            return default_ssid_2g_value

def wireless_change_ssid_name(driver, new_ssid):
    edit_2g_label = driver.find_elements_by_css_selector(label_name_in_2g)
    edit_2g_fields = driver.find_elements_by_css_selector(wrap_input)
    for l, f in zip(edit_2g_label, edit_2g_fields):
        # Connection type
        if l.text == 'Network Name(SSID)':
            for i in range(2):
                default_ssid_2g_value = f.find_element_by_css_selector(input)
                default_ssid_2g_value.clear()
                default_ssid_2g_value.send_keys(new_ssid)


def wireless_check_pw_eye(driver, block, change_pw=False, new_pw='00000000'):
    if not change_pw:
        pw_eye = block.find_element_by_css_selector(password_eye)
        act = ActionChains(driver)
        act.click_and_hold(pw_eye)
        pw_default = block.find_element_by_css_selector(input_pw).get_attribute('value')
        act.release(pw_eye)
        act.perform()
        return pw_default
    else:
        # Change password
        pw_place = block.find_element_by_css_selector(input_pw)
        ActionChains(driver).move_to_element(pw_place).click().key_down(Keys.CONTROL).send_keys('a').key_up(
            Keys.CONTROL).send_keys(new_pw).perform()
        return new_pw


def wireless_change_choose_option(driver, element_option, VALUE_OPTION):
    """
        Ap dung cho:
            Security
            Encryption
            Key Type
    """
    action_wl = driver.find_element_by_css_selector(element_option)
    action_wl.click()
    ls_options = action_wl.find_elements_by_css_selector(secure_value_in_drop_down)
    time.sleep(0.5)
    for o in ls_options:
        if o.get_attribute('option-value') == VALUE_OPTION:
            o.click()
            break


def nw_add_reserved_ip(driver, mac_add_value, ip_value):
    reserved_ip_block = driver.find_element_by_css_selector(network_reserved_ip_card)
    # Click Add
    reserved_ip_block.find_element_by_css_selector(add_class).click()
    edit_field = reserved_ip_block.find_element_by_css_selector(edit_mode)

    ip_addr_value = edit_field.find_element_by_css_selector(wol_mac_addr)
    ActionChains(driver).move_to_element(ip_addr_value).click().perform()
    time.sleep(0.5)

    reserved_ip_block.find_elements_by_css_selector(secure_value_in_drop_down)[-1].click()
    time.sleep(0.5)

    fill_mac_addr = edit_field.find_element_by_css_selector(wol_mac_addr)
    ActionChains(driver).click(fill_mac_addr).send_keys(mac_add_value).perform()

    ip_reserve_input = edit_field.find_element_by_css_selector(ele_reserve_ip_addr_input)
    ActionChains(driver).click(ip_reserve_input).send_keys(Keys.BACK_SPACE * 3).perform()
    time.sleep(1)
    IP_VALUE2 = ip_value.split('.')[-1]
    ActionChains(driver).click(ip_reserve_input).send_keys(IP_VALUE2).perform()


def check_enable_ethernet():
    import subprocess
    interface = subprocess.check_output('ipconfig', shell=True)
    if 'Ethernet adapter Ethernet:' not in interface.decode('utf8'):
        os.system(f'python {nw_interface_path} -i Ethernet -a enable')
        time.sleep(13)
    os.system(f'netsh wlan disconnect')
    time.sleep(2)


def wait_ethernet_available():
    count = 0
    while True:
        time.sleep(1)
        count += 1
        res = get_value_from_ipconfig('Ethernet adapter Ethernet', 'IPv4 Address')
        if res != 'Block or field error.':
            return True
        if (count % 180) == 0:
            return False


def wait_wifi_available():
    count = 0
    while True:
        time.sleep(1)
        count += 1
        res = get_value_from_ipconfig('Wireless LAN adapter Wi-Fi', 'IPv4 Address')
        if res != 'Block or field error.':
            return True
        if (count % 20) == 0:
            return False


def goto_system(driver, element_option):
    driver.find_element_by_css_selector(system_btn).click()
    time.sleep(0.2)
    driver.find_element_by_css_selector(element_option).click()
    time.sleep(1)


def choose_specific_value_from_dropdown(driver, dropdown_box_element, specific_value):
    driver.find_element_by_css_selector(dropdown_box_element).click()
    time.sleep(0.5)
    _options = driver.find_elements_by_css_selector(secure_value_in_drop_down)
    for o in _options:
        if o.text == specific_value:
            o.click()
            break


def choose_specific_radio_box(driver, LABEL, check=True):
    labels = driver.find_elements_by_css_selector(label_name_in_2g)
    values = driver.find_elements_by_css_selector(wrap_input)
    for l, v in zip(labels, values):
        if l.text.upper() == LABEL.upper():
            current_radio_status = v.find_element_by_css_selector(input)
            if check:
                if not current_radio_status.is_selected():
                    v.find_element_by_css_selector(select).click()
            else:
                if current_radio_status.is_selected():
                    v.find_element_by_css_selector(select).click()
            break


def connect_wifi(wifi_ssid, password):
    import os
    try:
        import pywifi
    except:
        os.system('pip install pywifi')
        import pywifi
    from pywifi import const
    os.system('netsh wlan disconnect')
    wifi = pywifi.PyWiFi()
    iface = wifi.interfaces()[0]
    iface.disconnect()
    time.sleep(1)

    profile = pywifi.Profile()
    profile.ssid = wifi_ssid
    profile.auth = const.AUTH_ALG_OPEN
    profile.akm.append(const.AKM_TYPE_WPA2PSK)
    profile.cipher = const.CIPHER_TYPE_CCMP
    profile.key = password

    # iface.remove_all_network_profiles()
    tmp_profile = iface.add_network_profile(profile)
    iface.connect(tmp_profile)
    time.sleep(12)


def current_connected_wifi():
    import subprocess
    ifaces = subprocess.check_output('netsh wlan show interfaces')
    if 'Profile' in ifaces.decode('utf8'):
        return ifaces.decode('utf8').split('Profile                :')[1].split('Hosted')[0].strip()
    return 'WiFi is not connected'

def get_current_wifi_MAC():
    import subprocess
    ifaces = subprocess.check_output('netsh wlan show interface mode=BSSID')
    lines = [i.strip() for i in ifaces.decode('utf8').splitlines()]
    mac_row = [i for i in lines if i.startswith('BSSID')]
    if len(mac_row):
        return mac_row[0].split('BSSID                  :')[1].strip().upper()
    return 'Wifi is not connected'

def factory_dut():
    save_config(config_path, 'URL', 'url', 'http://dearmyrouter.net')
    url_login = get_config('URL', 'url')
    filename_1 = '1'
    command = 'factorycfg.sh -a'
    run_cmd(command, filename_1)
    # Wait 5 mins for factory
    time.sleep(150)
    wait_DUT_activated(url_login)
    wait_ping('192.168.1.1')
    time.sleep(5)
    filename_2 = 'account2.txt'
    command_2 = 'capitest get Device.Users.User.2. leaf'
    run_cmd(command_2, filename_2)
    time.sleep(10)
    url_login = get_config('URL', 'url')
    # url = url_login + '/' + filename_2
    # _res = requests.get(url)
    # Get account information from web server and write to config.txt
    get_result_command_from_server_api(url_login, filename_2)


def call_api_login_old_firmware(user, pw):
    """Method: POST; Require: user, pw; Return: JSON data"""
    url = get_config('URL', 'url')
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encodev2(user, pw)
    }
    res = requests.post(url=url_login, json=data)

    if res.status_code != 200:

        user = get_config('ACCOUNT', 'user')
        pw = get_config('ACCOUNT', 'default_pw')
        data = {
            "userName": user,
            "password": base64encodev2(user, pw)
        }
        res = requests.post(url=url_login, json=data)
        if res.status_code == 200:
            save_config(config_path, 'ACCOUNT', 'password', pw)

    json_data = json.loads(res.text)
    return json_data


def detect_firmware_version(driver):
    url_login = get_config('URL', 'url')
    driver.get(url_login)
    time.sleep(1)
    if len(driver.find_elements_by_css_selector(lg_captcha_src)) == 0:
        user_request = get_config('ACCOUNT', 'user')
        pass_word = get_config('ACCOUNT', 'password')
        call_api_login_old_firmware(user_request, pass_word)
        user_request = get_config('ACCOUNT', 'user')
        pass_word = get_config('ACCOUNT', 'password')
        time.sleep(1)
        driver.get(url_login)
        time.sleep(2)
        driver.find_element_by_css_selector(el_lg_user_down_firm).send_keys(user_request)
        time.sleep(1)
        driver.find_element_by_css_selector(el_lg_pw_down_firm).send_keys(pass_word)
        time.sleep(1)
        driver.find_element_by_css_selector(el_lg_button_down_firm).click()
        time.sleep(5)
        policy_popup = driver.find_elements_by_css_selector('#privacy-policy-dialogue [name="privacyPolicyForm"]')
        if len(policy_popup) > 0:
            policy_popup[0].find_element_by_css_selector('p+p+p+p').location_once_scrolled_into_view
            time.sleep(1)
            driver.find_element_by_css_selector(ele_privacy_agree).click()
            time.sleep(5)
            driver.get(url_login+'/homepage')
            time.sleep(10)
        time.sleep(3)
        if len(driver.find_elements_by_css_selector('.fancybox-skin .new-version')) > 0:
            time.sleep(1)
            driver.find_element_by_css_selector('.fancybox-skin .btn-cancel').click()
            time.sleep(1)
        wait_visible(driver, el_home_wrap_down_firm)
        time.sleep(3)
        # ==============================================================================================
        driver.find_element_by_css_selector('#system_menu_popup').click()
        time.sleep(1)
        driver.find_element_by_css_selector('#wrap_system_menu [ng-click="openUpgradePopup()"]').click()
        time.sleep(3)
        os.chdir(files_path)
        firmware_30012_path = os.path.join(os.getcwd(), 't10x_fullimage_3.00.12_rev11.img')
        os.chdir(test_t10x_path)
        time.sleep(1)
        driver.find_element_by_css_selector('#upload-form-upgrade [uploader="uploader"]').send_keys(
            firmware_30012_path)
        time.sleep(1)
        driver.find_element_by_css_selector('#upload-form-upgrade [ng-click="applyUpgradePopup()"]').click()

        time.sleep(1)
        if len(driver.find_elements_by_css_selector('.custom-radio:nth-child(2) span')) > 0:
            driver.find_element_by_css_selector('.custom-radio:nth-child(2) span').click()
        time.sleep(0.5)
        driver.find_element_by_css_selector('.fancybox-opened [ng-click="confirmChange()"]').click()
        time.sleep(100)
        wait_visible(driver, '.fancybox-opened [ng-click="$parent.completedOk()"]')
        time.sleep(1)
        driver.find_element_by_css_selector('.fancybox-opened [ng-click="$parent.completedOk()"]').click()
        wait_ethernet_available()
        wait_ethernet_available()


def scan_wifi_repeater_mode(driver, wifi_name, wifi_pw):
    _rows = driver.find_elements_by_css_selector(rows)
    # Choose Network name
    for r in _rows:
        if r.find_element_by_css_selector(ele_network_name).text.strip() == wifi_name:
            r.click()
            time.sleep(1)
            break
    # Fill Password
    pw_box = driver.find_element_by_css_selector(ele_input_pw)
    ActionChains(driver).click(pw_box).send_keys(wifi_pw).perform()
    time.sleep(1)
    # Apply
    driver.find_element_by_css_selector(ele_apply_highlight).click()
    time.sleep(0.5)

    driver.find_element_by_css_selector(btn_ok).click()
    time.sleep(50)
    wait_popup_disappear(driver, icon_loading)
    time.sleep(1)
    wait_popup_disappear(driver, dialog_loading)
    time.sleep(5)
    wait_visible(driver, lg_page)
    save_config(config_path, 'URL', 'url', 'http://dearmyextender.net')


def connect_repeater_mode(driver, REPEATER_UPPER='', PW='', force=False):
    if REPEATER_UPPER == '':
        REPEATER_UPPER = get_config('REPEATER', 'repeater_name', input_data_path)
    if PW == '':
        PW = get_config('REPEATER', 'repeater_pw', input_data_path)

    if (driver.find_element_by_css_selector(ele_repeater_mode_input).is_selected() is False and force is False) or (force is True):
        driver.find_element_by_css_selector(ele_select_repeater_mode).click()
        time.sleep(0.5)
        driver.find_element_by_css_selector(apply).click()
        time.sleep(0.5)
        wait_popup_disappear(driver, dialog_loading)
        _rows = driver.find_elements_by_css_selector(rows)
        # Choose Network name
        for r in _rows:
            if r.find_element_by_css_selector(ele_network_name).text.strip() == REPEATER_UPPER:
                r.click()
                break
        # Fill Password
        pw_box = driver.find_element_by_css_selector(ele_input_pw)
        ActionChains(driver).click(pw_box).send_keys(PW).perform()
        time.sleep(1)
        # Apply
        driver.find_element_by_css_selector(ele_apply_highlight).click()
        time.sleep(0.5)
        driver.find_element_by_css_selector(btn_ok).click()
        time.sleep(1)
        wait_popup_disappear(driver, icon_loading)
        time.sleep(1)
        wait_popup_disappear(driver, icon_loading)
        wait_ethernet_available()
    save_config(config_path, 'URL', 'url', 'http://dearmyextender.net')



def connect_repeater_mode_third_party(driver, UPPER='Repeater_Upper_2G', PW='88888888'):
    if not driver.find_element_by_css_selector(ele_repeater_mode_input).is_selected():
        driver.find_element_by_css_selector(ele_select_repeater_mode).click()
        time.sleep(0.5)
        driver.find_element_by_css_selector(apply).click()
        time.sleep(0.5)
    else:
        goto_menu(driver, wireless_tab, wireless_repeater_setting_tab)
        wait_popup_disappear(driver, dialog_loading)

    wait_popup_disappear(driver, dialog_loading)
    _rows = driver.find_elements_by_css_selector(rows)
    # Choose Network name
    for r in _rows:
        if r.find_element_by_css_selector(ele_network_name).text.strip() == UPPER:
            r.click()
            time.sleep(1)
            break
    # Fill Password
    pw_box = driver.find_element_by_css_selector(ele_input_pw)
    ActionChains(driver).click(pw_box).send_keys(PW).perform()
    time.sleep(1)
    # Apply
    driver.find_element_by_css_selector(ele_apply_highlight).click()
    time.sleep(0.5)
    driver.find_element_by_css_selector(ele_apply_highlight).click()
    time.sleep(0.5)
    driver.find_element_by_css_selector(btn_ok).click()
    time.sleep(50)
    wait_popup_disappear(driver, icon_loading)
    time.sleep(1)
    wait_popup_disappear(driver, icon_loading)
    # wait_popup_disappear(driver, lg_page)
    save_config(config_path, 'URL', 'url', 'http://dearmyextender.net')

def change_firmware_version(driver, version='t10x_fullimage_3.00.12_rev11.img'):
    driver.find_element_by_css_selector(system_btn).click()
    time.sleep(1)
    driver.find_element_by_css_selector(ele_sys_firmware_update).click()
    time.sleep(1)
    os.chdir(files_path)
    firmware_path = os.path.join(os.getcwd(), version)
    driver.find_element_by_css_selector(ele_choose_firmware_file).send_keys(firmware_path)
    os.chdir(test_t10x_path)
    driver.find_element_by_css_selector(apply).click()
    time.sleep(1)
    if len(driver.find_elements_by_css_selector(ele_choose_firmware_select)) > 0:
        driver.find_element_by_css_selector(ele_choose_firmware_select).click()
    time.sleep(0.5)
    driver.find_element_by_css_selector(btn_ok).click()
    time.sleep(0.5)
    if len(driver.find_elements_by_css_selector(btn_ok)) > 0:
        driver.find_element_by_css_selector(btn_ok).click()
    time.sleep(150)

    wait_popup_disappear(driver, icon_loading)
    wait_visible(driver, content)
    wait_ethernet_available()
    driver.find_element_by_css_selector(btn_ok).click()
    time.sleep(1)

def check_connect_to_google():
    driver2 = webdriver.Chrome(driver_path)
    GOOGLE_URL = get_config('COMMON', 'google_url', input_data_path)
    driver2.get(GOOGLE_URL)
    time.sleep(5)
    check_google = len(driver2.find_elements_by_css_selector(google_img)) > 0
    driver2.quit()
    return check_google

def check_connect_to_youtube():
    driver2 = webdriver.Chrome(driver_path)
    YOUTUBE_URL = get_config('COMMON', 'youtube_url', input_data_path)
    driver2.get(YOUTUBE_URL)
    time.sleep(5)
    check_youtube = len(driver2.find_elements_by_css_selector('#logo-icon-container')) > 0
    driver2.quit()
    return check_youtube

def check_connect_to_web_admin_page():
    driver2 = webdriver.Chrome(driver_path)
    URL = get_config('URL', 'url')
    driver2.get(URL)
    time.sleep(5)
    check_web = len(driver2.find_elements_by_css_selector(lg_page)) > 0
    driver2.quit()
    return check_web


def disconnect_or_connect_wan(disconnected=True, URL_LOGIN='', _USER='', _PW=''):
    if URL_LOGIN == '':
        URL_LOGIN = get_config('URL', 'url')
    if _USER == '':
        _USER = get_config('ACCOUNT', 'user')
    if _PW == '':
        _PW = get_config('ACCOUNT', 'password')

    if disconnected:
        _URL = URL_LOGIN + '/api/v1/network/wan/0/disconnect'
    else:
        _URL = URL_LOGIN + '/api/v1/network/wan/0/connect'
    _METHOD = 'POST'
    _TOKEN = get_token(_USER, _PW)
    _BODY = ''
    call_api(_URL, _METHOD, _BODY, _TOKEN)
    time.sleep(10)


def add_port_forwarding(driver, SERVICE_TYPE,
                        IP_ADDRESS_SPLIT, LOCAL_START_END,
                        EXTERNAL_START_END, PROTOCOL_TYPE='TCP'):
    # Click Add button to change setting
    driver.find_element_by_css_selector(add_class).click()
    time.sleep(1)
    edit_field = driver.find_element_by_css_selector(edit_mode)
    time.sleep(0.5)
    # Fill Service type
    service_type = edit_field.find_element_by_css_selector(service_type_cls)
    service_type.find_element_by_css_selector(input).send_keys(SERVICE_TYPE)
    # IP address
    for i in range(2):
        ip_address = edit_field.find_element_by_css_selector(ip_address_col_cls)
        ip_address_box = ip_address.find_element_by_css_selector(input)
        ip_address_box.clear()
        ip_address_box.send_keys(IP_ADDRESS_SPLIT)
    # Local Port
    local_port = edit_field.find_element_by_css_selector(local_port_cls)
    local_port_input = local_port.find_elements_by_css_selector(input)
    for i, v in zip(local_port_input, LOCAL_START_END):
        i.clear()
        i.send_keys(v)
    # External Port
    external_port = edit_field.find_element_by_css_selector(external_port_cls)
    external_port_input = external_port.find_elements_by_css_selector(input)
    for i, v in zip(external_port_input, EXTERNAL_START_END):
        i.clear()
        i.send_keys(v)
    # Protocol
    protocol_box = edit_field.find_element_by_css_selector(protocol_col_cls)
    protocol_box.find_element_by_css_selector(option_select).click()
    time.sleep(0.2)
    ls_option = driver.find_elements_by_css_selector(active_drop_down_values)
    for o in ls_option:
        if o.text == PROTOCOL_TYPE:
            o.click()
            time.sleep(1)
            break

def get_port_forwarding_table(driver):
    table_value = list()
    ls_rows = driver.find_elements_by_css_selector(rows)
    if len(ls_rows):
        for r in ls_rows:
            row_active = r.find_element_by_css_selector(input).is_selected()
            row_service = r.find_element_by_css_selector(service_type_cls).text
            row_ip = r.find_element_by_css_selector(ip_address_col_cls).text
            row_local = r.find_element_by_css_selector(local_port_cls).text
            row_external = r.find_element_by_css_selector(external_port_cls).text
            row_protocol = r.find_element_by_css_selector(protocol_col_cls).text
            row_values = [row_active, row_service, row_ip, row_local, row_external, row_protocol]
            table_value.append(row_values)
    return table_value


def add_a_parental_control_rule(driver):
    rule_block = driver.find_element_by_css_selector(parental_rule_card)
    # Click Add 1
    rule_block.find_element_by_css_selector(add_class).click()
    time.sleep(1)
    rule_block = driver.find_element_by_css_selector(parental_rule_card)
    # Edit mode
    edit_field = rule_block.find_element_by_css_selector(edit_mode)

    device_name_field = edit_field.find_element_by_css_selector(name_cls)
    device_name_field.find_element_by_css_selector(input).click()

    # Select all
    opts = device_name_field.find_elements_by_css_selector(secure_value_in_drop_down)
    for i in range(len(opts) - 1):
        opts = device_name_field.find_elements_by_css_selector(secure_value_in_drop_down)
        opts[0].click()
        break
    time.sleep(1)
    # Setup the filter
    edit_field.find_element_by_css_selector('.service-filter').find_element_by_css_selector(apply).click()
    time.sleep(1)

    ls_service = driver.find_elements_by_css_selector('.service-item-wrap')
    for f in ls_service:
        if f.text == 'Social Network':
            f.click()
            break
    time.sleep(1)
    ls_service_sub = driver.find_elements_by_css_selector('.service-sub-item-wrap')
    for s in ls_service_sub:
        if s.text == 'facebook':
            if not len(s.find_elements_by_css_selector('.selected-icon')) > 0:
                s.click()
                break
    time.sleep(1)
    driver.find_element_by_css_selector(btn_ok).click()
    time.sleep(1)
    driver.find_element_by_css_selector(btn_save).click()
    wait_popup_disappear(driver, dialog_loading)
    time.sleep(2)


def add_a_ip_port_filtering(driver, DESC_VALUE, IP_ADDRESS_SPLIT, PORT_START_END, PROTOCOL_TYPE):
    filter_block = driver.find_element_by_css_selector(ele_ip_port_filtering)
    # Click Add 1
    filter_block.find_element_by_css_selector(add_class).click()
    time.sleep(1)
    filter_block = driver.find_element_by_css_selector(ele_ip_port_filtering)
    # Edit mode
    edit_field = filter_block.find_element_by_css_selector(edit_mode)

    # Fill Service type
    description_box = edit_field.find_element_by_css_selector(description)
    description_box.find_element_by_css_selector(input).send_keys(DESC_VALUE)
    # IP address
    ip_address = edit_field.find_element_by_css_selector(ip_address_cls)
    ip_address_box = ip_address.find_element_by_css_selector(input)
    for i in range(2):
        ip_address_box.clear()
        ip_address_box.send_keys(IP_ADDRESS_SPLIT)
    # Port Start End

    # for i in range(2):
    #     for i, v in zip(port_input, PORT_START_END):
    #         i.clear()
    #         i.send_keys(v)
    time.sleep(0.5)
    port_box = edit_field.find_element_by_css_selector(ele_port)
    port_input = port_box.find_elements_by_css_selector(input)


    for i in port_input:
        # for i in range(2):
        i.clear()
        i.clear()
        i.send_keys(PORT_START_END[0])
        time.sleep(0.2)

    time.sleep(0.5)
    # Protocol
    protocol_box = edit_field.find_element_by_css_selector(ele_protocol)
    protocol_box.find_element_by_css_selector(option_select).click()
    time.sleep(0.2)
    ls_option = driver.find_elements_by_css_selector(active_drop_down_values)
    for o in ls_option:
        if o.text == PROTOCOL_TYPE:
            o.click()
            time.sleep(1)
            break

    driver.find_element_by_css_selector(btn_save).click()
    time.sleep(2)
    driver.find_element_by_css_selector(apply).click()
    wait_popup_disappear(driver, dialog_loading)
    driver.find_element_by_css_selector(btn_ok).click()
    time.sleep(0.5)


def get_ip_port_filtering_table(driver):
    table_value = list()
    ls_rows = driver.find_elements_by_css_selector(rows)
    if len(ls_rows):
        for r in ls_rows:
            row_active = r.find_element_by_css_selector(input).is_selected()
            row_desc = r.find_element_by_css_selector(description).text
            row_ip = r.find_element_by_css_selector(ip_address_cls).text
            row_port = r.find_element_by_css_selector(ele_port).text
            row_protocol = r.find_element_by_css_selector(ele_protocol).text
            row_values = [row_active, row_desc, row_ip, row_port, row_protocol]
            table_value.append(row_values)
    return table_value


def connect_wifi_by_command(wifi_name, wifi_pw, xml_file=wifi_default_file_path):
    if xml_file == wifi_none_secure_path:
        write_data_to_none_secure_xml(xml_file, new_name=wifi_name)
    else:
        write_data_to_xml(xml_file, new_name=wifi_name, new_pw=wifi_pw)
    time.sleep(1)
    os.system(f'netsh wlan delete profile name="{wifi_name}"')
    time.sleep(0.5)
    # Connect Default 2GHz
    os.system(f'netsh wlan add profile filename="{xml_file}"')
    time.sleep(0.5)
    os.system(f'netsh wlan connect ssid="{wifi_name}" name="{wifi_name}"')
    wait_wifi_available()
    return current_connected_wifi()


def add_a_wireless_mac_filtering(driver, INPUT_DEVICE, INPUT_MAC):
    mac_block = driver.find_element_by_css_selector(ele_block_card)
    # Click Add 1
    mac_block.find_element_by_css_selector(add_class).click()
    time.sleep(1)
    mac_block = driver.find_element_by_css_selector(ele_block_card)
    # Edit mode
    edit_field = mac_block.find_element_by_css_selector(edit_mode)

    device_field = edit_field.find_element_by_css_selector(ele_mac_device_name)
    device_field.find_element_by_css_selector(input).send_keys(INPUT_DEVICE)

    # Select all
    mac_field = edit_field.find_element_by_css_selector(wol_mac_addr)
    mac_field.find_element_by_css_selector(input).click()

    CORRECT_OTHER_MAC = INPUT_MAC.replace(':', '')
    driver.find_element_by_css_selector('.user-define').click()
    time.sleep(0.5)
    driver.find_element_by_css_selector('.mac-address input').send_keys(CORRECT_OTHER_MAC)
    time.sleep(0.5)

    if driver.find_element_by_css_selector(btn_save).is_enabled():
        driver.find_element_by_css_selector(btn_save).click()
        time.sleep(0.5)


def add_a_mac_filtering(driver, OTHER_MAC=''):
    mac_block = driver.find_element_by_css_selector(ele_mac_filtering)
    # Click Add 1
    mac_block.find_element_by_css_selector(add_class).click()
    time.sleep(1)
    mac_block = driver.find_element_by_css_selector(ele_mac_filtering)
    # Edit mode
    edit_field = mac_block.find_element_by_css_selector(edit_mode)

    # Select all
    mac_field = edit_field.find_element_by_css_selector(wol_mac_addr)
    mac_field.find_element_by_css_selector(input).click()
    if OTHER_MAC == '':
        opts = mac_field.find_elements_by_css_selector(secure_value_in_drop_down)
        for i in range(len(opts) - 1):
            opts = mac_field.find_elements_by_css_selector(secure_value_in_drop_down)
            opts[0].click()
            break
        time.sleep(1)
        time.sleep(2)
        driver.find_element_by_css_selector(btn_save).click()
    else:
        CORRECT_OTHER_MAC = OTHER_MAC.replace(':', '')
        driver.find_element_by_css_selector('.user-define').click()
        time.sleep(0.5)
        driver.find_element_by_css_selector('.mac-address input').send_keys(CORRECT_OTHER_MAC)
        time.sleep(0.5)
        if driver.find_element_by_css_selector(btn_save).is_enabled():
            driver.find_element_by_css_selector(btn_save).click()
            time.sleep(0.5)
        else:
            driver.find_element_by_css_selector('.input-cancel-button').click()

            edit_field = driver.find_element_by_css_selector(edit_mode)
            mac_field = edit_field.find_element_by_css_selector(wol_mac_addr)
            mac_field.find_element_by_css_selector(input).click()
            time.sleep(0.5)
            driver.find_element_by_css_selector('.user-define').click()
            time.sleep(0.5)
            tmp_mac = random_mac_address().replace(':', '')
            driver.find_element_by_css_selector('.mac-address input').send_keys(tmp_mac)
            driver.find_element_by_css_selector(btn_save).click()
            time.sleep(0.5)


    time.sleep(1)
    driver.find_element_by_css_selector(apply).click()
    time.sleep(0.5)
    driver.find_element_by_css_selector(btn_ok).click()
    wait_popup_disappear(driver, dialog_loading)
    driver.find_element_by_css_selector(btn_ok).click()
    wait_popup_disappear(driver, dialog_loading)

def get_mac_filtering_table(driver):
    mac_block = driver.find_element_by_css_selector(ele_mac_filtering)
    table_value = list()
    ls_rows = mac_block.find_elements_by_css_selector(rows)
    if len(ls_rows):
        for r in ls_rows:
            row_active = r.find_element_by_css_selector(input).is_selected()
            row_device_name = r.find_element_by_css_selector(ele_mac_device_name).text
            row_mac = r.find_element_by_css_selector(wol_mac_addr).text
            row_values = [row_active, row_device_name, row_mac]
            table_value.append(row_values)
    return table_value


def random_mac_address():
    import random
    ls_random = ['0', '2', '4', '6', '8', 'A', 'C', 'E']
    mac = [
        random.choices(ls_random, k=2),
        random.choices(ls_random, k=2),
        random.choices(ls_random, k=2),
        random.choices(ls_random, k=2),
        random.choices(ls_random, k=2),
        random.choices(ls_random, k=2)]
    return ':'.join([''.join(i) for i in mac]).upper()


def change_password(driver, CURRENT_PW, NEW_PW):
    fields = driver.find_elements_by_css_selector(ele_input_pw)
    fields[0].send_keys(CURRENT_PW)
    fields[1].send_keys(NEW_PW)
    fields[2].send_keys(NEW_PW)


def add_a_port_triggering(driver,
                          DESC_VALUE,
                          TRIGGERED_START_END,
                          PROTOCOL_TYPE_TRIGGERED,
                          FORWARDED_START_END,
                          PROTOCOL_TYPE_FORWARDED):
    triggering_block = driver.find_element_by_css_selector(port_triggering_card)
    # Click Add 1
    triggering_block.find_element_by_css_selector(add_class).click()
    time.sleep(1)
    triggering_block = driver.find_element_by_css_selector(port_triggering_card)
    # Edit mode
    edit_field = triggering_block.find_element_by_css_selector(edit_mode)

    # Fill Service type
    description_box = edit_field.find_element_by_css_selector(description_col_cls)
    description_box.find_element_by_css_selector(input).send_keys(DESC_VALUE)

    # Port Start End
    trigger_range = edit_field.find_element_by_css_selector(triggered_col_cls)
    trigger_range.find_elements_by_css_selector(input)[0].send_keys(TRIGGERED_START_END[0])
    trigger_range.find_elements_by_css_selector(input)[1].send_keys(TRIGGERED_START_END[1])
    # Protocol
    protocol_box_trigger = edit_field.find_elements_by_css_selector(protocol_col_cls)[0]
    protocol_box_trigger.find_element_by_css_selector(option_select).click()
    time.sleep(0.2)
    ls_option = driver.find_elements_by_css_selector(active_drop_down_values)
    for o in ls_option:
        if o.text == PROTOCOL_TYPE_TRIGGERED:
            o.click()
            time.sleep(1)
            break
    # Port Start End
    forwarding_range = edit_field.find_element_by_css_selector(forwarded_col_cls)
    forwarding_range.find_elements_by_css_selector(input)[0].send_keys(FORWARDED_START_END[0])
    forwarding_range.find_elements_by_css_selector(input)[1].send_keys(FORWARDED_START_END[1])
    time.sleep(0.5)
    # Protocol
    protocol_box_forwarding = edit_field.find_elements_by_css_selector(protocol_col_cls)[1]
    protocol_box_forwarding.find_element_by_css_selector(option_select).click()
    time.sleep(0.2)
    ls_option = driver.find_elements_by_css_selector(active_drop_down_values)
    for o in ls_option:
        if o.text == PROTOCOL_TYPE_FORWARDED:
            o.click()
            time.sleep(1)
            break
    driver.find_element_by_css_selector(btn_save).click()
    time.sleep(2)
    driver.find_element_by_css_selector(apply).click()
    wait_popup_disappear(driver, dialog_loading)
    time.sleep(1)
    driver.find_element_by_css_selector(btn_ok).click()
    time.sleep(1)


def add_a_usb_network_folder(driver, DESC_VALUE, PATH_FILE, WRITE=True):
    network_block = driver.find_element_by_css_selector(usb_network)
    network_block.find_element_by_css_selector(add_class).click()

    # Edit mode
    network_block = driver.find_element_by_css_selector(usb_network)
    edit_field = network_block.find_element_by_css_selector(edit_mode)
    # Description
    description_field = edit_field.find_element_by_css_selector(description)
    description_field.find_element_by_css_selector(input).send_keys(DESC_VALUE)
    # Folder path
    path_field = edit_field.find_element_by_css_selector(path)
    path_field.find_element_by_css_selector(input).click()
    time.sleep(0.5)
    # Choose path
    driver.find_element_by_css_selector(tree_icon).click()
    time.sleep(0.2)

    ls_path_lv1 = driver.find_elements_by_css_selector(path_name_lv1)
    for o in ls_path_lv1:
        if o.text == PATH_FILE:
            ActionChains(driver).move_to_element(o).click().perform()
            break
    # OK
    driver.find_element_by_css_selector(btn_ok).click()
    time.sleep(1)
    if WRITE:
        edit_field.find_element_by_css_selector('.read-write #custom-checkbox-write-add+label').click()
    else:
        edit_field.find_element_by_css_selector('.read-write #custom-checkbox-read-add+label').click()
    time.sleep(1)
    network_block.find_element_by_css_selector(btn_save).click()
    time.sleep(1)
    network_block.find_element_by_css_selector(apply).click()
    wait_popup_disappear(driver, dialog_loading)
    time.sleep(1)
    driver.find_element_by_css_selector(btn_ok).click()
    wait_popup_disappear(driver, dialog_loading)
    time.sleep(1)


def add_a_usb_account_setting(driver, ID_VALUE, PASSWORD_VALUE):
    account_settings_block = driver.find_element_by_css_selector(account_setting_card)
    account_settings_block.find_element_by_css_selector(add_class).click()

    # Edit mode
    account_settings_block = driver.find_element_by_css_selector(account_setting_card)
    edit_field = account_settings_block.find_element_by_css_selector(edit_mode)
    # ID
    id_field = edit_field.find_element_by_css_selector(id_cls)
    id_field.find_element_by_css_selector(input).send_keys(ID_VALUE)
    # PASSWORD
    pw_field = edit_field.find_element_by_css_selector(password_cls)
    pw_field.find_element_by_css_selector(input).send_keys(PASSWORD_VALUE)

    account_settings_block.find_element_by_css_selector(btn_save).click()
    time.sleep(1)
    account_settings_block.find_element_by_css_selector(apply).click()
    wait_popup_disappear(driver, dialog_loading)
    time.sleep(1)
    driver.find_element_by_css_selector(btn_ok).click()
    wait_popup_disappear(driver, dialog_loading)
    time.sleep(1)


def add_a_default_guest_network(driver, block):
    block.find_element_by_css_selector(add_class).click()
    time.sleep(0.5)
    # Check Default Value
    edit_2g_block = driver.find_elements_by_css_selector(wl_primary_card)[0]
    # Settings
    wl_2g_ssid = wireless_get_default_ssid(edit_2g_block, 'Network Name(SSID)')
    # Click Hide SSID
    edit_2g_block.find_elements_by_css_selector(select)[0].click()
    time.sleep(0.5)
    confirm_msg_2g = driver.find_element_by_css_selector(confirm_dialog_msg).text
    time.sleep(0.5)
    driver.find_element_by_css_selector(btn_ok).click()
    # Apply
    edit_2g_block.find_element_by_css_selector(apply).click()
    wait_popup_disappear(driver, dialog_loading)
    time.sleep(0.5)
    return wl_2g_ssid


def add_a_full_guest_network(driver, block, SECURE, ENCRYPT='', KEY_TYPE='', _PW=''):
    block.find_element_by_css_selector(add_class).click()
    time.sleep(0.5)
    # Check Default Value
    edit_block = driver.find_elements_by_css_selector(wl_primary_card)[0]
    # Settings
    time.sleep(0.5)
    wl_ssid = wireless_get_default_ssid(edit_block, 'Network Name(SSID)')
    time.sleep(0.5)
    if SECURE == 'NONE':
        wireless_change_choose_option(edit_block, secure_value_field, SECURE)
        wifi_pw = ''
        time.sleep(0.5)
    elif SECURE == 'WPA2-PSK':
        wireless_change_choose_option(edit_block, secure_value_field, SECURE)
        time.sleep(0.5)
        wifi_pw = wireless_check_pw_eye(driver, edit_block, change_pw=False)
        time.sleep(0.5)
    elif SECURE == 'WPA2/WPA-PSK':
        wireless_change_choose_option(edit_block, secure_value_field, SECURE)
        if ENCRYPT != '':
            wireless_change_choose_option(edit_block, encryption_value_field, ENCRYPT)
            time.sleep(0.5)
        if _PW != '':
            wireless_check_pw_eye(driver, edit_block, change_pw=True, new_pw=_PW)
        time.sleep(0.5)
        wifi_pw = wireless_check_pw_eye(driver, edit_block, change_pw=False)
        time.sleep(0.5)
    else:
        wireless_change_choose_option(edit_block, secure_value_field, SECURE)
        time.sleep(0.5)
        if ENCRYPT != '':
            wireless_change_choose_option(edit_block, encryption_value_field, ENCRYPT)
            time.sleep(0.5)
        if KEY_TYPE != '':
            wireless_change_choose_option(edit_block, key_type_value_field, KEY_TYPE)
            time.sleep(0.5)

        wifi_pw = wireless_check_pw_eye(driver, edit_block, change_pw=True, new_pw=_PW)
        time.sleep(0.5)

    edit_block.find_element_by_css_selector(apply).click()
    wait_popup_disappear(driver, dialog_loading)
    time.sleep(0.5)
    return {"name": wl_ssid, "password": wifi_pw}


def scan_wifi_repeater_mode_table(driver):
    ls_row = driver.find_elements_by_css_selector(rows)
    table = list()
    for r in ls_row:
        row_data = [i.text for i in r.find_elements_by_css_selector('.col')]
        table.append(row_data)
    return table


def interface_connect_disconnect(interface, _status):
    """
    :param interface: Ethernet or Wi-Fi
    :param _status: enable or disable
    :return:
    """
    # Correct params
    if interface.lower() == 'ethernet':
        interface = 'Ethernet'
    elif interface.lower() in ['wifi', 'wi-fi']:
        interface = 'Wi-Fi'
    if _status.lower() in ['disable', 'disabled']:
        _status = 'disable'
        sleep_time = 5
    elif _status.lower() in ['enable', 'enabled']:
        _status = 'enable'
        sleep_time = 10

    os.system(f'python {nw_interface_path} -i {interface} -a {_status}')
    time.sleep(sleep_time)


# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~Driver serve  BLOCK~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
def find_element(driver_or_parent, element_locator, locator_type=By.CSS_SELECTOR, time_out_s=5) -> WebElement:
    """
    Find element from driver. If not found element after time out, raise time out error
    :param driver_or_parent: selenium webdriver or parent element.
    :param element_locator: locator - used to find the element.
    :param locator_type: type of locator. Default is CSS.
    :param time_out_s: time out in second to find element. Default is 5 seconds.
    :return: WebElement
    """
    if driver_or_parent is None or (
            not isinstance(driver_or_parent, WebDriver) and not isinstance(driver_or_parent, WebElement)):
        raise TypeError("Driver must be not none and is instance of selenium webdriver or is instance of selenium webelement")
    list_valid_locator_types = [
        By.XPATH,
        By.CSS_SELECTOR,
        By.ID,
        By.NAME,
        By.CLASS_NAME
    ]
    string_to_method = {
        By.XPATH: driver_or_parent.find_element_by_xpath,
        By.CSS_SELECTOR: driver_or_parent.find_element_by_css_selector,
        By.ID: driver_or_parent.find_element_by_id,
        By.NAME: driver_or_parent.find_element_by_name,
        By.CLASS_NAME: driver_or_parent.find_element_by_class_name
    }
    if locator_type is None or locator_type not in list_valid_locator_types:
        raise TypeError(f"Locator type must be not none and in list {str(list_valid_locator_types)}.")

    time_stamp_s = 0.05
    while time_out_s >= 0:
        try:
            return string_to_method[locator_type](element_locator)
        except:
            time_out_s = time_out_s - time_stamp_s
            time.sleep(time_stamp_s)
    raise TimeoutError(f"Element define by '{element_locator}' is not presence after {time_out_s} seconds.")


def find_elements(driver_or_parent, element_locator, locator_type=By.CSS_SELECTOR, time_out_s=5):
    """
    Find list elements from driver or parent element. If not found element after time out, raise time out error
    :param driver_or_parent: selenium webdriver or parent element.
    :param element_locator: locator - used to find the element.
    :param locator_type: type of locator. Default is CSS.
    :param time_out_s: time out in second to find element. Default is 5 seconds.
    :return: list of WebElement
    """
    if driver_or_parent is None or (
            not isinstance(driver_or_parent, WebDriver) and not isinstance(driver_or_parent, WebElement)):
        raise TypeError("Driver must be not none and is instance of selenium webdriver or is instance of selenium webelement")
    list_valid_locator_types = [
        By.XPATH,
        By.CSS_SELECTOR,
        By.ID,
        By.NAME,
        By.CLASS_NAME
    ]
    string_to_method = {
        By.XPATH: driver_or_parent.find_elements_by_xpath,
        By.CSS_SELECTOR: driver_or_parent.find_elements_by_css_selector,
        By.ID: driver_or_parent.find_elements_by_id,
        By.NAME: driver_or_parent.find_elements_by_name,
        By.CLASS_NAME: driver_or_parent.find_elements_by_class_name
    }
    if locator_type is None or locator_type not in list_valid_locator_types:
        raise TypeError(f"Locator type must be not none and in list {str(list_valid_locator_types)}.")

    time_stamp_s = 0.05
    while time_out_s >= 0:
        try:
            return string_to_method[locator_type](element_locator)
        except:
            time_out_s = time_out_s - time_stamp_s
            time.sleep(time_stamp_s)
    raise TimeoutError(f"Element define by '{element_locator}' is not presence after {time_out_s} seconds.")


def click_to_element(driver_or_parent, element_locator, locator_type=By.CSS_SELECTOR, time_out_s=5):
    """
    Clicks the element.
    :param driver_or_parent: selenium webdriver or parent element.
    :param element_locator: locator - used to find the element.
    :param locator_type: type of locator. Default is CSS.
    :param time_out_s: time out in second to find element. Default is 5 seconds.
    :return:
    """
    element = find_element(driver_or_parent, element_locator, locator_type, time_out_s)
    element.click()


def is_element_displayed(driver_or_parent, element_locator, locator_type=By.CSS_SELECTOR, time_out_s=5) -> bool:
    """
    Whether the element is visible to a user.
    :param driver_or_parent: selenium webdriver or parent web element.
    :param element_locator: locator - used to find the element.
    :param locator_type: type of locator. Default is CSS.
    :param time_out_s: time out in second to find element. Default is 5 seconds.
    :return:
    """
    try:
        element = find_element(driver_or_parent, element_locator, locator_type, time_out_s)
        return element.is_displayed()
    except:
        return False

def get_part_from_key(key: str = None):
    list_valid_part = ["MAIN", "HOME", "NETWORK", "WIRELESS", "SECURITY", "ADVANCED", "MEDIASHARE", "NON_FUNCTION"]
    for part in list_valid_part:
        if key.startswith(part):
            return part
    raise TypeError(f"Not valid key name. Key must start with one of following text: {','.join(list_valid_part)}")


def save_duration_time(test_case_key, test_case_name, test_case_steps, start_time):
    from datetime import datetime
    if "[END TC]" in str(test_case_steps):
        next_duration = (datetime.now() - start_time)
        part = get_part_from_key(test_case_key)
        str_current_duration = get_config(section=part, option=test_case_name, config_dir=testcase_runtime_data_path)
        if str_current_duration is None or str_current_duration == '0':
            duration_s = next_duration.total_seconds()
        else:
            current_duration = int(str_current_duration)
            duration_s = (int(current_duration) + next_duration.total_seconds())/2
        save_config(testcase_runtime_data_path,
                    part,
                    test_case_name,
                    str(round(duration_s+0.5)))


def convert_stage_of_step_to_string(value):
    """
    Convert stage of step to string
    :param value:
    :return:
    """
    bool_to_string = {
        True: "[Pass]",
        False: "[Fail]",
    }
    return bool_to_string[value]


def generate_step_information(step_name, list_check_in_step, list_actual, list_expected):
    step_status = assert_list(list_actual,list_expected)["result"]
    step_info = f"{convert_stage_of_step_to_string(step_status)} {step_name}\n"
    for i in range(0, len(list_check_in_step)):
        step_info = f"{step_info}\t - {list_check_in_step[i]}\n"

    step_info = f"{step_info}\tActual:\n"
    for i in range(0, len(list_actual)):
        if isinstance(list_actual[i], bool):
            step_info = f"{step_info}\t - {detect_check_information(list_check_in_step[i], list_actual[i]==list_expected[i])}\n"
        else:
            step_info = f"{step_info}\t - {list_actual[i]}\n"

    step_info = f"{step_info}\tExpected:\n"
    for i in range(0, len(list_expected)):
        if isinstance(list_actual[i], bool):
            step_info = f"{step_info}\t - {detect_check_information(list_check_in_step[i], True)}\n"
        else:
            step_info = f"{step_info}\t - {list_expected[i]}\n"

    return step_info


def detect_check_information(checking_info: str = None, result: bool = None) -> str:
    dict_opposite_stage = {
        "on": "off",
        "off": "on",
        "check": "uncheck",
        "enabled": "disabled",
        "disabled": "enabled",
        "appear": "not appear",
        "success": "unsuccess",
        "unsuccess": "success",
        "connect": "not connect",
        "not connect": "connect",
        "contain:": "not contain:",
        "not contain:": "contain:",
        "kept": "not kept",
        "not kept": "kep",
        "is displayed": "is not displayed",
        "is not displayed": "is displayed",
        "existed": "not existed",
        "not existed": "existed",
        "correct": "not correct",
        "not correct": "correct"
    }
    for key in dict_opposite_stage:
        if checking_info.endswith(key) or \
                ("contain:" in key and key in checking_info) or \
                (("is displayed:" == key or "is not displayed:" == key) and key in checking_info):
            if result:
                return checking_info
            else:
                index = checking_info.rfind(key)
                return f"{checking_info[:index]} {dict_opposite_stage[key]} {checking_info[index+len(key):]}"
    raise TypeError(f"Not define action for step: {checking_info}")
