import os
import time
import pyodbc as pyodbc
from PIL import Image
from io import BytesIO
import requests
from requests.exceptions import Timeout
import configparser
import json
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
import inspect
from openpyxl.styles import Border, Side
import datetime
import shutil
import base64
from path import root_dir

def save_config(config_path, section, option, value):
    if not os.path.exists(config_path):
        print("The config file not exist. Exit!!!")
        return

    config = configparser.RawConfigParser()
    config.read(config_path)

    if not config.has_section(str(section).upper()):
        config.add_section(str(section).upper())

    if (str(get_config(section, option)) != str(value)):
        config.set(str(section).upper(), str(option), str(value))
        with open(config_path, 'w') as config_file:
            config.write(config_file)

def get_config(config_path, section, option):
    if not os.path.exists(config_path):
        print("The config file not exist. Exit!!!")
        return

    config = configparser.RawConfigParser()
    config.read(config_path)

    if config.has_option(str(section).upper(), option):
        return config.get(str(section).upper(), option)
    else:
        return

ifconfig_path = os.path.join(root_dir, "Config", "ifconfig.txt")
offical_path = os.path.join(root_dir, "Config", "WEB_UI", "offical_report.xlsx")
final_path = os.path.join(root_dir, "Config", "WEB_UI", "final_report.xlsx")
official_api_path = os.path.join(root_dir, "Config", "WEB_API", "official_api_report.xlsx")
final_api_path = os.path.join(root_dir, "Config", "WEB_API", "final_api_report.xlsx")
# from HGJ310.HGJ310_Webui import final_report
url = 'http://' + str(get_config(ifconfig_path, "IFCONFIG", "ipv4"))
user = get_config(ifconfig_path, "USER_INFO", "user")
pass_word = get_config(ifconfig_path, "USER_INFO", "pw")
rg_port = get_config(ifconfig_path, "PORT", "rg_port")
import logging

logging.basicConfig(filename='example.log', format='%(asctime)s,%(msecs)d %(levelname)-8s [%(filename)s:%(lineno)d] %(message)s',
    datefmt='%Y-%m-%d:%H:%M:%S',
    level=logging.DEBUG)

def get_index_page():
    api_url = url + "/index.html"
    try:
        res = requests.get(url=api_url, timeout=1)
        if (res.status_code == 200):
            return True
    except Exception:
        return False

def wait_DUT_activated():
    count = 0
    start = time.time()
    while True:
        count += 1
        if get_index_page():
            end = time.time()
            break
        else:
            print("      WAITING FOR DUT ACTIVED... | RE-TRY TIMES: " + str(count))
    duration = end - start
    return duration

def get_web_key():
    url_req = url + '/api/v1/gateway/users/login/auth'
    res = requests.get(url_req)
    if res.status_code == 200:
        salt = json.loads(res.text)['web_key']
        return salt
    else:
        return None

def base64encode(user, pw):
    import hashlib, binascii, base64
    salt = get_web_key()
    if salt is not None:
        mode = 'sha512'
        iterations = 2048
        dklen = 32
        hash0_pw = hashlib.pbkdf2_hmac(mode, pw.encode('utf-8'), salt.encode('utf-8'), iterations, dklen)
        hash1_pw = binascii.hexlify(hash0_pw)
        hash2_pw = hash1_pw.decode('utf-8')
        initial_data = 'HS:' + user + ':' + hash2_pw
        byte_string = initial_data.encode('utf-8')
        encoded_data = base64.b64encode(byte_string)
        return encoded_data.decode('ascii')
    else:
        initial_data = 'HS:' + user + ':' + pw
        byte_string = initial_data.encode('utf-8')
        encoded_data = base64.b64encode(byte_string)

        return encoded_data.decode('ascii')

def get_func_name():
    return inspect.stack()[1][3]

def thread(func_name):
    t = open('func_name.txt', 'a')
    t.write(func_name + '\n')
    t.close()
    os.system('func_name.txt')


def reset_report_result(report):
    report_xlsx = offical_path
    wb = openpyxl.load_workbook(report_xlsx)
    ws = wb.active
    for i in range(2, ws.max_row + 1):
        ws.cell(row=i, column=5).value = None
        ws.cell(row=i, column=6).value = None
        ws.cell(row=i, column=7).value = None
    wb.save(report_xlsx)

    final_report_root = final_path
    wb = openpyxl.load_workbook(final_report_root)
    ws = wb.active
    for i in range(2, ws.max_row + 1):
        ws.delete_rows(2)
    wb.save(final_report_root)

    shutil.copy(final_report_root, report)


def write_actual_excel(list_steps, func_name, duration, final_report):
    report_xlsx = offical_path
    str_steps = ''
    for i in list_steps:
        str_steps += (i + '\n')
    wb = openpyxl.load_workbook(report_xlsx)
    ws = wb.active
    # Write list steps
    for i in range(2, ws.max_row + 1):
        if func_name == ws.cell(i, 1).value:
            logging.debug("")
            # Write to final rp
            row_value = []
            for j in range(2, ws.max_column + 1):
                row_value.append(ws.cell(row=i, column=j).value)
                logging.debug("")
            fi_wb = openpyxl.load_workbook(final_report)
            fi_ws = fi_wb.active
            r = 1
            while True:
                if fi_ws.cell(row=r, column=2).value is None:
                    for c in range(2, fi_ws.max_column+1):
                        fi_ws.cell(row=r, column=c).value = row_value[c-2]
                        fi_ws.cell(r, 7).value = str_steps
                        fi_ws.cell(r, 6).value = duration
                        fi_ws.cell(r, 1).value = str(fi_ws.max_row-1)
                        logging.debug("")
                    break
                else:
                    r += 1
                    continue

            fi_wb.save(final_report)
            paint_result(final_report)
            # Write test steps
            ws.cell(i, 7).value = str_steps
            # Write duration
            ws.cell(i, 6).value = duration

            break
    wb.save(report_xlsx)
    paint_result(report_xlsx)

    # Hide unnecessary row
    wb = openpyxl.load_workbook(report_xlsx)
    ws = wb.active
    for i in range(2, ws.max_row+1):
        ws.row_dimensions[i].hidden = False
        if ws.cell(row=i, column=7).value is None:
            logging.debug("")
            ws.row_dimensions[i].hidden = True
            logging.debug("")
    wb.save(report_xlsx)


def set_border(ws, cell_range):
    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

    rows = ws.iter_rows(cell_range)
    for row in rows:
        for cell in row:
            cell.border = border


def paint_result(report_xlsx):
    redFill = PatternFill(start_color='FF0000',
                          end_color='FF0000',
                          fill_type='solid')
    GreenFill = PatternFill(start_color='32CD32',
                            end_color='32CD32',
                            fill_type='solid')
    OrangeFill = PatternFill(start_color='FF8000',
                             end_color='FF8000',
                             fill_type='solid')
    wb = openpyxl.load_workbook(report_xlsx)
    ws = wb.active
    for i in range(2, ws.max_row + 1):
        if ws.cell(row=i, column=7).value is not None:
            if '[Fail]' in ws.cell(row=i, column=7).value:
                ws.cell(row=i, column=5).value = 'NG'
                ws.cell(row=i, column=5).fill = OrangeFill
            else:
                if '[Pass]' not in ws.cell(row=i, column=7).value or ws.cell(row=i, column=7).value is None:
                    ws.cell(row=i, column=5).value = 'ER'
                    ws.cell(row=i, column=5).fill = redFill
                else:
                    ws.cell(row=i, column=5).value = 'OK'
                    ws.cell(row=i, column=5).fill = GreenFill
    wb.save(report_xlsx)


def convert_rbg_hex(rgb):
    a = rgb.split('(')[1].split(')')[0].split(',')
    return '#{:02x}{:02x}{:02x}'.format(int(a[0]), int(a[1]), int(a[2]))


def check_radio_tick(driver, css_element):
    logging.debug("")
    result = driver.find_element_by_css_selector(css_element).get_property('checked')
    logging.debug("")
    return result


def login(driver, self, url):
    logging.debug("")
    check_login(driver, self, url)
    logging.debug("")
    time.sleep(1)
    logging.debug("")
    driver.execute_script('''document.title="''' + self._testMethodName + '''"''')
    logging.debug("")
    driver.find_element_by_id('login').send_keys(user)
    logging.debug("")
    driver.find_element_by_id('senha').send_keys(pass_word)
    logging.debug("")
    driver.find_element_by_xpath('//button[@value="Entrar"]').click()
    logging.debug("")
    time.sleep(1)


def test_login(driver, self, url, user, pass_word):
    time.sleep(1)
    driver.get(url)
    driver.execute_script('''document.title="''' + self._testMethodName + '''"''')
    time.sleep(1)
    lg_box = driver.find_element_by_id('login')
    ActionChains(driver).move_to_element(lg_box).click().key_down(Keys.CONTROL).send_keys('a').key_up(
        Keys.CONTROL).send_keys(Keys.DELETE).send_keys(user).perform()

    pw_box = driver.find_element_by_id('senha')
    ActionChains(driver).move_to_element(pw_box).click().key_down(Keys.CONTROL).send_keys('a').key_up(
        Keys.CONTROL).send_keys(Keys.DELETE).send_keys(pass_word).perform()
    driver.find_element_by_xpath('//button[@value="Entrar"]').click()
    time.sleep(1)


def check_login(driver, self, url):
    # Log in
    count = 0
    while 1:
        count += 1
        driver.get(url)
        logging.debug("")
        time.sleep(3)
        check = len(driver.find_elements_by_id('login'))
        logging.debug("")
        driver.execute_script('''document.title="''' + self._testMethodName + '''"''')
        logging.debug("")
        if check != 0:
            logging.debug("")
            break
        else:
            if count == 10:
                logging.debug("")
                lg_box = driver.find_elements_by_id('login')
                logging.debug("")
                if len(lg_box) == 0:
                    logging.debug("")
                    os.system(
                        '''"C:\Program Files\VanDyke Software\SecureCRT\SecureCRT.exe" /SCRIPT ./reboot.py /SERIAL '''
                        + rg_port + ''' /BAUD 115200''')
                    logging.debug("")
                    while 1:
                        driver.get(url)
                        logging.debug("")
                        time.sleep(0.5)
                        check = len(driver.find_elements_by_id('login'))
                        logging.debug("")
                        if check != 0:
                            logging.debug("")
                            break


def get_wenui_components(driver):
    api = []
    for request in driver.requests:
        logging.debug("")
        if request.path != 'http://nosrc.fbiz.com.br/270x600':
            api.append([request.path, request.response.reason])
        logging.debug("")

    for i in api:
        check_url = i[0].split('//')[1].split('/')[0]
        logging.debug("")
        if check_url == 'www.gstatic.com' or check_url == 'accounts.google.com':
            logging.debug("")
            api.remove(i)
            logging.debug("")
    a = []
    for i in api:
        if i[0].split('/')[-1] == '':
            logging.debug("")
            a.append([i[0].split('/')[-2], i[1]])
            logging.debug("")
        else:
            logging.debug("")
            a.append([i[0].split('/')[-1], i[1]])
            logging.debug("")
    return a


def check_current_request(driver, request_path):
    global body
    import ast
    # Reverse this list request to get the api latest
    list_request = driver.requests
    list_request.reverse()
    for request in list_request:
        if request_path in request.path:
            body = request.response.body
            dict_str = ast.literal_eval(body.decode('utf-8'))
            break
        else:
            dict_str = "Request Not Found"

    return dict_str


def wait_time(self, driver):
    logging.debug("")
    pop_up_wait = driver.find_elements_by_css_selector('.msgText')
    logging.debug("")
    count_time = 0
    logging.debug("")
    while len(pop_up_wait) == 1:
        logging.debug("")
        pop_up_wait = driver.find_elements_by_css_selector('.msgText')
        time.sleep(1)
        logging.debug("")
        count_time += 1
        logging.debug("")
        if count_time >= 300:
            logging.debug("")
            break
        logging.debug("")
    self.assertLessEqual(count_time, 300, 'Pop-up was timeout > 5 minutes')


def wait_visible(driver, css_element):
    visible = driver.find_elements_by_css_selector(css_element)
    count = 0
    while not len(visible):
        time.sleep(1)
        count += 1
        visible = driver.find_elements_by_css_selector(css_element)
        if count == 30:
            driver.refresh()
        if count == 100:
            driver.quit()


def call_get_api_genaeral(url_req):
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + url_req
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~


def call_put_api(url_req, token, body=None):
    """Method: PUT; Require: Token"""
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    data = body
    res = requests.put(url=url_req, headers=headers, json=data, timeout=30)
    json_data = json.loads(res.text)
    return json_data, res.status_code


def call_put_api_extend(url_req, headers, body=None):
    """Method: PUT; Require: Headers"""
    data = body
    res = requests.put(url=url_req, headers=headers, json=data, timeout=30)
    json_data = json.loads(res.text)
    return json_data, res.status_code


def call_post_api(url_req, token, body=None):
    """Method: PÓT; Require: Token"""
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    data = body
    res = requests.post(url=url_req, headers=headers, json=data, timeout=30)
    json_data = json.loads(res.text)
    return json_data, res.status_code


def call_post_api_extend(url_req, headers, body=None):
    """Method: POST; Require: headers"""
    data = body
    res = requests.post(url=url_req, headers=headers, json=data, timeout=30)#
    json_data = json.loads(res.text)
    return json_data, res.status_code


def call_api_login(user, pw):
    """Method: POST; Require: user, pw; Return: JSON data"""
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res = requests.post(url=url_login, json=data)
    json_data = json.loads(res.text)
    return json_data, res.status_code


def call_api_login_extend(url, data):
    res = requests.post(url=url, json=data)
    json_data = json.loads(res.text)
    return json_data, res.status_code


def call_api_logout(token):
    url_logout = url + "/api/v1/gateway/users/logout"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.post(url=url_logout, headers=headers)
    json_data = json.loads(res.text)
    return json_data, res.status_code


def call_get_api_token(url_req, accesstoken):
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": accesstoken
    }
    res = requests.get(url=url_req, headers=headers)
    json_data = json.loads(res.text)
    return json_data, res.status_code


def call_get_api_extend(url_req, headers):
    res = requests.get(url=url_req, headers=headers)
    json_data = json.loads(res.text)
    return json_data, res.status_code


def report_excel_api(list_steps, func_name, final_api_report):
    report_xlsx = official_api_path
    str_steps = ''
    for i in list_steps:
        str_steps += (i + '\n')
    wb = openpyxl.load_workbook(report_xlsx)
    ws = wb.active
    # Write list steps
    for i in range(2, ws.max_row + 1):
        if func_name == ws.cell(i, 1).value:
            # Write to final rp
            row_value = []
            for j in range(2, ws.max_column + 1):
                row_value.append(ws.cell(row=i, column=j).value)
            fi_wb = openpyxl.load_workbook(final_api_report)
            fi_ws = fi_wb.active
            r = 1
            while True:
                if fi_ws.cell(row=r, column=2).value is None:
                    for c in range(2, fi_ws.max_column+1):
                        fi_ws.cell(row=r, column=c).value = row_value[c-2]
                        fi_ws.cell(r, 6).value = str_steps
                        fi_ws.cell(r, 1).value = str(fi_ws.max_row-1)

                    break
                else:
                    r += 1
                    continue

            fi_wb.save(final_api_report)
            paint_api_result(final_api_report)
            # Write test steps
            ws.cell(i, 6).value = str_steps

            break
    wb.save(report_xlsx)
    paint_api_result(report_xlsx)

    # Hide unnecessary row
    wb = openpyxl.load_workbook(report_xlsx)
    ws = wb.active
    for i in range(2, ws.max_row+1):
        ws.row_dimensions[i].hidden = False
        if ws.cell(row=i, column=7).value is None:
            ws.row_dimensions[i].hidden = True
    wb.save(report_xlsx)


def reset_report_api_result(report):
    report_xlsx = official_api_path
    wb = openpyxl.load_workbook(report_xlsx)
    ws = wb.active
    for i in range(2, ws.max_row + 1):
        ws.cell(row=i, column=5).value = None
        ws.cell(row=i, column=6).value = None
    wb.save(report_xlsx)

    final_report_root = final_api_path
    wb = openpyxl.load_workbook(final_report_root)
    ws = wb.active
    for i in range(2, ws.max_row + 1):
        ws.delete_rows(2)
    wb.save(final_report_root)

    shutil.copy(final_report_root, report)


def paint_api_result(report_xlsx):
    redFill = PatternFill(start_color='FF0000',
                          end_color='FF0000',
                          fill_type='solid')
    GreenFill = PatternFill(start_color='32CD32',
                            end_color='32CD32',
                            fill_type='solid')
    OrangeFill = PatternFill(start_color='FF8000',
                             end_color='FF8000',
                             fill_type='solid')
    wb = openpyxl.load_workbook(report_xlsx)
    ws = wb.active
    for i in range(2, ws.max_row + 1):
        if ws.cell(row=i, column=6).value is not None:
            if '[Fail]' in ws.cell(row=i, column=6).value:
                ws.cell(row=i, column=5).value = 'NG'
                ws.cell(row=i, column=5).fill = OrangeFill
            else:
                if '[Pass]' not in ws.cell(row=i, column=6).value or ws.cell(row=i, column=6).value is None:
                    ws.cell(row=i, column=5).value = 'ER'
                    ws.cell(row=i, column=5).fill = redFill
                else:
                    ws.cell(row=i, column=5).value = 'OK'
                    ws.cell(row=i, column=5).fill = GreenFill
    wb.save(report_xlsx)


def tag(*tags):
    """
    Decorator to add tags to a test class or method.
    """
    def decorator(obj):
        setattr(obj, 'tags', set(tags))
        return obj
    return decorator


# ~~~~~~~~~~~~~~~~~~~~~ From Security ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

def read_input(tc_name, step):
    with open('data_input', encoding='utf8') as p:
        step_input = json.load(p)
    step_i = step_input[tc_name][step]
    url = step_i['url']
    method = step_i['method']
    body = step_i['body']
    if body is None:
        return {"url": url, "method": method}
    else:
        return {"url": url, "method": method, "body": body}


def flatten_json(y):
    out = {}

    def flatten(x, name=''):
        if type(x) is dict:
            for a in x:
                flatten(x[a], name + a + '_')
        elif type(x) is list:
            i = 0
            for a in x:
                flatten(a, name + str(i) + '_')
                i += 1
        else:
            out[name[:-1]] = x
    flatten(y)
    return out


def assert_value(expected_tc, expected_step, res_actual):
    with open('data_expected', encoding='utf8') as de:
        data_expected = json.load(de)

        flat_expected = flatten_json(data_expected[expected_tc][expected_step])
        flat_actual = flatten_json(res_actual)
        actual = []
        expected = []
        for e in flat_expected.items():
            check_key = False
            for a in flat_actual.items():
                if e[0] == a[0]:
                    check_key = True
                    if e[1] == "_NOT_NULL_":
                        if len(str(a[1])) == 0:
                            expected.append(e)
                            actual.append(a)
                    elif e[1] == "_NULL_":
                        if len(str(a[1])) != 0:
                            expected.append(e)
                            actual.append(a)
                    break
            if not check_key:
                expected.append("Not existed " + str(e))
    if len(actual) == 0 and len(expected) == 0:
        return True, actual, expected
    return False, actual, expected


def assert_completely(expected_tc, expected_step, res_actual):
    with open('data_expected', encoding='utf8') as e:
        data_expected = json.load(e)
        flat_expected = flatten_json(data_expected[expected_tc][expected_step])
        flat_actual = flatten_json(res_actual)
    actual = []
    expected = []
    for e in flat_expected.items():
        check_key = False
        for a in flat_actual.items():
            if e[0] == a[0]:
                check_key = True
                if e[1] != a[1]:
                    expected.append(e)
                    actual.append(a)
                break
        if not check_key:
            expected.append("Not existed " + str(e))
    if len(actual) == 0 and len(expected) == 0:
        return True, actual, expected
    return False, actual, expected


def call_api(url, method, token, body=None):
    global res
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    if method.upper() == 'GET':
        res = requests.get(url=url, headers=headers, timeout=30)
    elif method.upper() == 'POST':
        res = requests.post(url=url, headers=headers, json=body, timeout=30)
    elif method.upper() == 'PUT':
        res = requests.put(url=url, headers=headers, json=body, timeout=30)
    elif method.upper() == 'DELETE':
        res = requests.put(url=url, headers=headers, timeout=30)
    json_data = json.loads(res.text)
    return json_data, res.status_code


def call_api_extend(url, method, headers, body=None):
    global res
    if method.upper() == 'GET':
        res = requests.get(url=url, headers=headers, timeout=30)
    elif method.upper() == 'POST':
        res = requests.post(url=url, headers=headers, json=body, timeout=30)
    elif method.upper() == 'PUT':
        res = requests.put(url=url, headers=headers, json=body, timeout=30)
    elif method.upper() == 'DELETE':
        res = requests.put(url=url, headers=headers, timeout=30)
    json_data = json.loads(res.text)
    return json_data, res.status_code


# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
def api_login():

    # login and get token
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }

    res_login = requests.post(url=url_login, json=data)
    json_data = json.loads(res_login.text)
    return json_data


def test_api_login(user, pass_word):

    # login and get token
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }

    res_login = requests.post(url=url_login, json=data)
    json_data = json.loads(res_login.text)
    return json_data


def api_gateway_about():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/gateway/about"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)

    json_data_about = json.loads(res.text)
    return json_data_about


def api_network_docsis():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/network/docsis"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data_docsis = json.loads(res.text)
    return json_data_docsis


def api_network_docsis_initial_frequency():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/network/docsis/initialFrequency"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_network_docsis_procedure():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/network/docsis/procedure"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_network_docsis_channel():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/network/docsis/channel"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_network_wan():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/network/wan"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_mta_startup_procedure():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/network/docsis/mta/procedure"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_mta_line_status():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/network/docsis/mta/line"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_network_lan():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/network/lan"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_network_lan_restoreDefault():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/network/lan/ipv6/restoreDefault"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_lan_ipv6():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/gateway/devices?interface=lan-2.4g-5g"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def get_wan_info():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/network/wan"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def get_lan_info():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/network/lan"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def get_lan_v6_info():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/gateway/devices?interface=lan-2.4g-5g"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_dmz():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/dmz"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_ddns():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/ddns"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_networkOption():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/networkOption"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_upnp():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/upnp"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_firewall():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/security/firewall"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_ipv4_firewall():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/security/firewall/ipv4"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_firewall_alert():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/security/firewall/alert"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_service_log():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/log"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_service_port_ip():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/portIpFiltering"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_service_port_ip_delete_all():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/portIpFiltering/deletes"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    requests.post(url=url_request, headers=headers)


def api_service_mac_filtering():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/macFiltering"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_service_macfiltering_delete_all():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/macFiltering/deletes"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    requests.post(url=url_request, headers=headers)


def api_service_port_filtering():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/portFiltering"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_service_portfiltering_delete_all():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/portFiltering/deletes"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    requests.post(url=url_request, headers=headers)


def api_portForWarding():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/portForwarding"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_portTriggering():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/portTriggering"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_accessADM():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/accessADM"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_wifi_radio(id):

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/wifi/" + str(id) + "/radio"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_wifi_scanResult(id):

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/wifi/" + str(id) + "/scanResult"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_wifi_accessControl(id):

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/wifi/" + str(id) + "/ssid/0/accessControl"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_security_managedDevices():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/security/managedDevices"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_wifi_ssid_2G():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/wifi/0/ssid/0"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_wifi_ssid_5G():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/wifi/1/ssid/1"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_wifi_radio_2G():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/wifi/0/radio"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_wifi_radio_5G():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/wifi/1/radio"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_wifi_wps_2G():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/wifi/wps?interfaceId=0"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_wifi_wps_5G():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/wifi/wps?interfaceId=1"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_wifi_wds_2G():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/wifi/0/wds"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_wifi_wds_5G():

    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": base64encode(user, pass_word)
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/wifi/1/wds"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data
