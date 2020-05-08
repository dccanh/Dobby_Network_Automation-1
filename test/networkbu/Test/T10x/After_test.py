import sys, os
sys.path.append('../../')
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import datetime
from Helper.t10x.ls_path import gg_credential_path
a = str(datetime.datetime.strftime(datetime.datetime.now(), '%Y-%m-%d %H:%M'))

scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
creds = ServiceAccountCredentials.from_json_keyfile_name(gg_credential_path, scope)
client = gspread.authorize(creds)

sheet = client.open("[DOB] Report Automation").get_worksheet(0)
# Update Report Release Date
sheet.update_cell(7, 7, a)



import openpyxl
from Helper.t10x.common import get_config

get_gg_sheet_name = get_config('REPORT', 'sheet_name')

scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
creds = ServiceAccountCredentials.from_json_keyfile_name(gg_credential_path, scope)
client = gspread.authorize(creds)

sheet = client.open("[DOB] Report Automation").worksheet(get_gg_sheet_name)
table = sheet.get_all_values()
print(table)

unix_string = datetime.datetime.today().strftime('%Y-%m-%d_%H_%M_%S')
customize_report_path = os.path.join(get_config('REPORT', 'report_path'), f"Report_automation_{unix_string}.xlsx")

wb = openpyxl.load_workbook('excel_file.xlsx')
ws = wb.active
ws.delete_rows(6, ws.max_row)
for r in range(5, len(table)):
    for c in range(len(table[r])):
        ws.cell(row=r+1, column=c+1).value = table[r][c]
ws.cell(row=1, column=4).value = get_gg_sheet_name
wb.save(customize_report_path)