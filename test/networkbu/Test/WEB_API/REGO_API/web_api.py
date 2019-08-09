import os
import sys

sys.path.append('../../../')
import json
import time
import unittest
import datetime
import requests
import configparser
import Helper.Helper_common
from path import root_dir
from Helper.Helper_common import tag
from requests.exceptions import ReadTimeout
from passlib.hash import pbkdf2_sha512
import hashlib, binascii, base64
import openpyxl, shutil
from openpyxl.styles import PatternFill

class LOGGING(object):
    def __init__(self, *files):
        self.files = files

    def write(self, obj):
        for f in self.files:
            f.write(obj)

    def flush(self):
        pass

class WEB_API(object):
    def __init__(self):
        self.script_dir = os.path.dirname(os.path.realpath(__file__))
        self.log_path = str(os.path.join(self.script_dir, "test.log"))
        self.log_file = open(self.log_path, 'a')
        self.backup = sys.stdout
        sys.stdout = LOGGING(sys.stdout, self.log_file)

        self.time_stamp =  str(datetime.datetime.now())
        print('\n\n                     ---------------------------------------')
        print('==================== | RUNTIME: ' + self.time_stamp + ' | ====================')
        print('                     ---------------------------------------\n')

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        self.ifconfig_path = os.path.join(root_dir, "Config", "ifconfig.txt")
        self.offical_path = os.path.join(root_dir, "Config", "WEB_UI", "offical_report.xlsx")
        self.final_path = os.path.join(root_dir, "Config", "WEB_UI", "final_report.xlsx")
        self.official_api_path = os.path.join(root_dir, "Config", "WEB_API", "official_api_report.xlsx")
        self.final_api_path = os.path.join(root_dir, "Config", "WEB_API", "final_api_report.xlsx")
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        self.config = configparser.ConfigParser()
        self.config.read_file(open(self.ifconfig_path, 'r'))
        self.ipv4 = 'http://' + self.config.get("IFCONFIG", 'ipv4')
        self.user = self.config.get("USER_INFO", 'user')
        self.pw = self.config.get("USER_INFO", 'pw')
        self.new_pw = self.user
        self.precondition = False
        self.pre_step_res = None
        self.count = 0

        with open('data_input', encoding='utf8') as i:
            self.data_input = json.load(i)

        with open('data_expected', encoding='utf8') as e:
            self.data_expected = json.load(e)

        self.final_api_report = 'final_api_report_' + str(datetime.datetime.now()).replace(' ', '_').replace(':', '-') + '.xlsx'
        self.final_api_report = os.path.join(root_dir, "Report", "WEB_API", self.final_api_report)
        Helper.Helper_common.reset_report_api_result(self.final_api_report)

        self.session_timeout = 0

    def call_api(self, tc_name, step, token):
        self.res = {}
        step_i = self.data_input[tc_name][step]
        print("    ---> " + str(step))
        url = []
        if (type(step_i['url']) is str):
            url.append(step_i['url'])
        else:
            for i in range(0, len(step_i['url'])):
                url.append(str(step_i['url'][i]))
        # print(str(url))

        method = step_i['method']
        body = step_i['body']

        if "headers" in step_i.keys():
            headers = {
                "content-type": "application/json",
                "content-language": "en"
            }
            if 'access-token' in step_i['headers'].keys():
                item = {"access-token": str(step_i['headers']['access-token'].encode('utf-8'))}
                headers.update(item)
            else:
                item = step_i['headers']
                headers.update(item)
                # headers.update({str(list(item.keys())[0])+': ' + str(list(item.values())[0])})
        else:
            headers = {
                "content-type": "application/json",
                "content-language": "en",
                "access-token": self.token
            }

        body_keys = list(body.keys())
        if ("userName" in body_keys):
            if (body["userName"] == "_NULL_"):
                body["userName"] = ""
            elif (body["userName"] == "_USERNAME_"):
                body["userName"] = self.user

        if ("password" in body_keys):
            if (body["password"] == "_NULL_"):
                body["password"] = ""
            elif (body["password"] == "_HASH_PASSWORD_"):
                body["password"] = self.pw_hash_encode

        if ("oldPassword" in body_keys):
            body["oldPassword"] = self.pw_hash_encode

        if ("newPassword" in body_keys):
            body["newPassword"] = self.new_pw_hash_encode
        # print(str(json.dumps(body)))

        for i in range(0, len(url)):
            api_uri = str(self.ipv4 + url[i])
            # print(str(api_uri))

            res_data = self.send_request(api_uri, method, headers, body)
            # print(str(type(res_data)))
            # print(str(res_data.text))
            # print(str(res_data))
            if (res_data.status_code != 404):
                json_data = json.loads(res_data.text)
                # print(str(json_data))
                data = {"body": json_data, "statusCode": res_data.status_code}
                if ("GW_17" in str(tc_name)) and ("1" in str(step)):
                    json_data_keys = list(self.flatten_json(json_data).keys())
                    # print(str(json_data_keys))
                    for j in range(0, len(json_data_keys)):
                        # print(str(json_data_keys[i]))
                        if ("session" in str(json_data_keys[j])):
                            self.session_timeout = self.flatten_json(json_data)[json_data_keys[j]]
            else:
                data = {"body": "404", "statusCode": 404}

            self.res[url[i]] = data
        # print(str(json.dumps(res)))
        # return {"body": json_data, "statusCode": res.status_code}

        if ("NW_18" in str(tc_name)) and (("2" in str(step)) or ("4" in str(step))):
            self.ipv4 = 'http://' + body["ipv4"]["ipAddress"]
            # print(str(self.ipv4))
            self.sleep_sec(5)

        return self.res

    def send_request(self, url, method, headers, body, timeout = 120):
        # print("url: " + str(url))
        # print("method: " + str(method))
        # print("headers: " + str(headers))
        # print("body: " + str(body))

        result = requests.models.Response()
        try:
            if method.upper() == 'GET':
                result = requests.get(url=url, headers=headers, json=body, timeout=timeout)
            elif method.upper() == 'POST':
                result = requests.post(url=url, headers=headers, json=body, timeout=timeout)
            elif method.upper() == 'PUT':
                result = requests.put(url=url, headers=headers, json=body, timeout=timeout)
            elif method.upper() == 'DELETE':
                result = requests.put(url=url, headers=headers, json=body, timeout=timeout)
        except Exception:
            result.status_code = 404

        # print(str(result.text))
        # print(str(result.status_code))
        return result

    def flatten_json(self, y):
        out = {}

        def flatten(x, name=''):
            if type(x) is dict:
                for a in x:
                    flatten(x[a], name + a + '_')
            elif type(x) is list:
                i = 0
                for a in x:
                    flatten(a, name + str(i) + '_')
                    i += 1
            else:
                out[name[:-1]] = x

        flatten(y)
        return out


    def assert_result(self, expected_tc, expected_step, response):
        flat_expected_body = self.flatten_json(self.data_expected[expected_tc][expected_step]['body'])
        flat_actual_body = self.flatten_json(response['body'])
        for e in flat_expected_body.items():
            if (str(e[0]) == "_ANY_"):
                if (e[1] == "_SAVE_"):
                    self.pre_step_res = flat_actual_body
                flat_expected_body = self.pre_step_res
        # print("\nflat_expected_body: " + str(flat_expected_body))
        # print("\nflat_actual_body: " + str(flat_actual_body))
        # print("\nflat_expected_body.items(): " + str(flat_expected_body.items()))
        # print("\nflat_actual_body.items(): " + str(flat_actual_body.items()))

        actual = []
        expected = []

        for e in flat_expected_body.items():
            # print(str(e))
            check_key = False

            for a in flat_actual_body.items():
                if e[0] == a[0]:
                    # print(str(e[0]) + " = " + a[0])
                    check_key = True
                    if e[1] == "_NOT_NULL_":
                        if len(str(a[1])) == 0:
                            expected.append(e)
                            actual.append(a)
                    elif e[1] == "_NULL_":
                        if len(str(a[1])) != 0:
                            expected.append(e)
                            actual.append(a)
                    elif e[1] != a[1]:
                        if (len(str(e[1])) != 0):
                            expected.append(e)
                            actual.append(a)
                    break
            if not check_key:
                expected.append("Not existed " + str(e))
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        expected_code = self.data_expected[expected_tc][expected_step]['statusCode']
        actual_code = response['statusCode']
        if (expected_code == "_NOT_501_"):
            if str(actual_code) == "501":
                actual.append("Status Code is {} but expected is {}.".format(actual_code, expected_code))
        elif (expected_code == "_NOT_200_"):
            if str(actual_code) == "200":
                actual.append("Status Code is {} but expected is {}.".format(actual_code, expected_code))
        elif expected_code != actual_code:
            actual.append("Status Code is {} but expected is {}.".format(actual_code, expected_code))
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # print("\nactual: " + str(actual))
        # print("\nexpected: " + str(expected))
        if (len(actual) == 0 and len(expected) == 0) or (self.precondition):
            return True, actual, expected
        return False, actual, expected

    def report_excel_api(self, list_steps, func_name, final_api_report):
        report_xlsx = self.official_api_path
        str_steps = ''
        for i in list_steps:
            str_steps += (i + '\n')
        wb = openpyxl.load_workbook(report_xlsx)
        ws = wb.active
        # Write list steps
        for i in range(2, ws.max_row + 1):
            if func_name == ws.cell(i, 1).value:
                # Write to final rp
                row_value = []
                for j in range(2, ws.max_column + 1):
                    row_value.append(ws.cell(row=i, column=j).value)
                fi_wb = openpyxl.load_workbook(final_api_report)
                fi_ws = fi_wb.active
                r = 1
                while True:
                    if fi_ws.cell(row=r, column=2).value is None:
                        for c in range(2, fi_ws.max_column+1):
                            fi_ws.cell(row=r, column=c).value = row_value[c-2]
                            fi_ws.cell(r, 6).value = str_steps
                            fi_ws.cell(r, 1).value = str(fi_ws.max_row-1)

                        break
                    else:
                        r += 1
                        continue

                fi_wb.save(final_api_report)
                self.paint_api_result(final_api_report)
                # Write test steps
                ws.cell(i, 6).value = str_steps

                break
        wb.save(report_xlsx)
        self.paint_api_result(report_xlsx)

        # Hide unnecessary row
        wb = openpyxl.load_workbook(report_xlsx)
        ws = wb.active
        for i in range(2, ws.max_row+1):
            ws.row_dimensions[i].hidden = False
            if ws.cell(row=i, column=7).value is None:
                ws.row_dimensions[i].hidden = True
        wb.save(report_xlsx)

    def reset_report_api_result(self, report):
        report_xlsx = official_api_path
        wb = openpyxl.load_workbook(report_xlsx)
        ws = wb.active
        for i in range(2, ws.max_row + 1):
            ws.cell(row=i, column=5).value = None
            ws.cell(row=i, column=6).value = None
        wb.save(report_xlsx)

        final_report_root = final_api_path
        wb = openpyxl.load_workbook(final_report_root)
        ws = wb.active
        for i in range(2, ws.max_row + 1):
            ws.delete_rows(2)
        wb.save(final_report_root)

        shutil.copy(final_report_root, report)

    def paint_api_result(self, report_xlsx):
        redFill = PatternFill(start_color='FF0000',
                              end_color='FF0000',
                              fill_type='solid')
        GreenFill = PatternFill(start_color='32CD32',
                                end_color='32CD32',
                                fill_type='solid')
        OrangeFill = PatternFill(start_color='FF8000',
                                 end_color='FF8000',
                                 fill_type='solid')
        wb = openpyxl.load_workbook(report_xlsx)
        ws = wb.active
        for i in range(2, ws.max_row + 1):
            if ws.cell(row=i, column=6).value is not None:
                if '[Fail]' in ws.cell(row=i, column=6).value:
                    ws.cell(row=i, column=5).value = 'NG'
                    ws.cell(row=i, column=5).fill = OrangeFill
                else:
                    if '[Pass]' not in ws.cell(row=i, column=6).value or ws.cell(row=i, column=6).value is None:
                        ws.cell(row=i, column=5).value = 'ER'
                        ws.cell(row=i, column=5).fill = redFill
                    else:
                        ws.cell(row=i, column=5).value = 'OK'
                        ws.cell(row=i, column=5).fill = GreenFill
        wb.save(report_xlsx)

    def get_token(self):
        self.wait_DUT_activated()
        self.token = Helper.Helper_common.call_api_login(self.user, self.pw)[0]["accessToken"]
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        self.pw_hash_encode = Helper.Helper_common.base64encode(self.user, self.pw)
        self.new_pw_hash_encode = Helper.Helper_common.base64encode(self.user, self.new_pw)
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

    def implement_test(self, tc_name, generate_report=True):
        # print(str(tc_name))
        self.get_token()

        check_list = []
        step_keys = list(self.data_input[tc_name].keys())
        list_steps = []
        # print(str(step_keys))
        for step in step_keys:
            try:
                if ("GW_13" in str(tc_name)) and ("4" in str(step)):
                    self.set_pre_info("login_info")

                res = self.call_api(tc_name, step, self.token)
                # print("\nres: " + str(json.dumps(res)))
                res_keys = list(res.keys())
                # print(str(res_keys))

                for i in range(0, len(res.keys())):
                    check = self.assert_result(tc_name, step, res[res_keys[i]])
                    check_list.append(check[0])

                    if check[0]:
                        list_steps.append(
                            '\n[Pass] {}. Assertion pass.'.format(step))
                    else:
                        list_steps.append(
                            '\n[Fail] {}. Assertion fail. \nActual: {}.\nExpected: {}.'.format(step, str(check[1]), str(check[2])))

                if ("GW_17" in str(tc_name)) and ("2" in str(step)):
                    self.sleep_min(self.session_timeout)

                if ("NW_12" in str(tc_name)) and (("2" in str(step)) or ("3" in str(step))):
                    self.sleep_sec(30)
                    self.wait_DUT_activated()
                    self.get_token()

            except ReadTimeout:
                list_steps.append('\n[Fail] {}. Call API timeout. ')
        # print("\check_list:" + str(check_list))

        test_result = "PASS"
        for i in range(0,len(check_list)):
            if not check_list[i]:
                test_result = "_____FALSE_____"

        print("         -----> RESULT: " + test_result)

        if generate_report:
            self.report_excel_api(list_steps, tc_name, self.final_api_report)

    def sleep_min(self, minutes):
        print("      WAITING FOR " + str(minutes) + " MINUTES...")
        time.sleep(minutes * 60)

    def sleep_sec(self, seconds):
        print("      WAITING FOR " + str(seconds) + " SECONDS...")
        time.sleep(seconds)

    def set_pre_info(self, mode):
        pre_info_script = os.path.join(root_dir, "Helper", "pre_info.bat")
        # print(str(login_info_path))
        model = self.config.get("PORT", 'model')
        # print(str(model))
        rg_port = self.config.get("PORT", 'rg_port')
        # print(str(rg_port))
        cm_port = self.config.get("PORT", 'cm_port')
        # print(str(cm_port))
        f = open(pre_info_script,"r")
        contents = f.readlines()
        f.close()

        bak = list(contents)
        # print(str(type(contents)))
        # print(str(bak))

        for i in range(0, len(contents)):
            if "python run_start.py" in str(contents[i]):
                cmd = "python run_start.py " + model + " " + rg_port + " " + cm_port + " " + str(mode)
                # print(str(a[i]))
                contents[i] = cmd
                # print(str(a[i]))

        # print(str(contents))
        f = open(pre_info_script,"w")
        for i in range(0, len(contents)):
            f.write(contents[i])
        f.close()
        os.system(pre_info_script)

        f = open(pre_info_script,"w")
        for i in range(0, len(bak)):
            # print(str(bak[i]))
            f.write(bak[i])
        f.close()

        if ((str(mode) == "reboot") or (str(mode) == "restore_defaults")):
            self.sleep_sec(30)

    def get_test_list(self, list_file):
        test_list = []
        with open(list_file, encoding='utf8') as file:
            for line in file:
                if not line.strip(): continue
                if (str(line).strip()[0] != "#"):
                    test_list.append(str(line).strip())
        return test_list

    def auto_run_tc(self):
        # self.set_pre_info("login_info")

        count = 0
        # test_list = ['test_BE_SV_13', 'test_BE_SV_54', 'test_BE_SV_55', 'test_BE_SV_56', 'test_BE_GW_01']
        test_list = self.get_test_list("test_list")
        start = time.time()

        for i in range(0, len(test_list)):
            self.wait_DUT_activated()
            self.precondition = False
            count += 1
            time_stamp =  str(datetime.datetime.now())
            print("%03d" % count + " | " + time_stamp + " | " + str(test_list[i]))
            start_tc = time.time()

            # if (i > 0):
            #     pre_tc = test_list[i - 1][0:10]
            #     cur_tc = test_list[i][0:10]
            #     if (str(pre_tc) != str(cur_tc)):
            #         # print("     pre_tc: " + str(pre_tc) + " | cur_tc: " + str(cur_tc))
            #         print("     RESTORING THE FACTORY DEDAULTS...")
            #         self.set_pre_info("restore_defaults")

            precondition_tc_list = {
                    "BE_SV_57": ["BE_SV_49"],
                    "BE_SV_68": ["BE_SV_62"],
                    "BE_SV_70": ["BE_SV_35", "BE_SV_32"],
                    "BE_SV_71": ["BE_SV_35", "BE_SV_32"],
                    "BE_SV_72": ["BE_SV_42", "BE_SV_39"],
                    "BE_SV_73": ["BE_SV_42", "BE_SV_39"],
                    "BE_SV_74": ["BE_SV_47", "BE_SV_44"],
                    "BE_SV_75": ["BE_SV_47", "BE_SV_44"],
                    "BE_SV_76": ["BE_SV_68", "BE_SV_62"],
                    "BE_SV_90": ["BE_SV_42"],
                    "BE_SV_91": ["BE_SV_42", "BE_SV_39"],
                    "BE_SV_96": ["BE_SV_47"],
                    "BE_SV_97": ["BE_SV_47", "BE_SV_44"],
                    "BE_SV_100": ["BE_SV_57", "BE_SV_49"],
                    "BE_SV_102": ["BE_SV_57", "BE_SV_49"],
                    "BE_SV_103": ["BE_SV_57", "BE_SV_49"],
                    "BE_WF_30": ["BE_WF_16"],
                    "BE_WF_31": ["BE_WF_16"],
                    "BE_WF_33": ["BE_WF_16"],
                    "BE_WF_34": ["BE_WF_16"],
                    "BE_WF_36": ["BE_WF_16"],
                    "BE_WF_37": ["BE_WF_16"]
                }

            precondition_tc_keys = list(precondition_tc_list.keys())
            precondition_tc = []
            # print(str(precondition_tc_keys))
            for key in precondition_tc_keys:
                if key in test_list[i]:
                    print("     SETTING UP THE PRE-CONDITIONS...")
                    precondition_tc = precondition_tc_list[key]
            for tc in precondition_tc:
                pre_tc = "test_" + tc
                print("           " + str(pre_tc))
                self.precondition = True
                self.implement_test(pre_tc, False)

            self.implement_test(test_list[i])
            end_tc = time.time()
            duration_tc = end_tc - start_tc
            print("         THE DURATION OF TEST CASE: " + "%10d" % duration_tc + " SECONDS.")

        end = time.time()
        self.duration = end - start
        print("\nTHE TOTAL DURATION TIME: " + "%10d" % self.duration + " SECONDS.")

    def get_index_page(self):
        api_url = self.ipv4 + "/index.html"
        try:
            res = requests.get(url=api_url, timeout=1)
            if (res.status_code == 200):
                return True
        except Exception:
            return False

    def wait_DUT_activated(self):
        self.count = 0
        while True:
            self.count += 1
            if self.get_index_page():
                break
            else:
                print("          WAITING FOR THE DEVICE ACTIVED... | RE-TRY TIMES: " + str(self.count))
                if (self.count % 180 == 0):
                    print("     REBOOTING THE DEVICE...")
                    self.set_pre_info("reboot")

api = WEB_API()
api.auto_run_tc()
