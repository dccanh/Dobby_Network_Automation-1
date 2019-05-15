import configparser
import inspect
import json
import openpyxl
import os
import pyodbc as pyodbc
import requests
import sys
import time
import utils.common

from PIL import Image
from io import BytesIO
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

script_dir = os.path.dirname(os.path.realpath(__file__))
config_file = str(os.path.join(script_dir, "..", "..", "config", "config.ini"))
if not os.path.exists(config_file):
    print("The config_file: " + config_file + "not exist. Exit!!!")
    sys.exit()

config = configparser.RawConfigParser()
config.read(config_file)

url = 'http://' + config.get('IP', 'ipv4')
user = config.get('AUTHENTICATION', 'gw_user')
pass_word = config.get('AUTHENTICATION', 'gw_pw')
report_excel = config.get('COMMON', 'report_xlsx')


def querydb(query):
    conn = pyodbc.connect('DRIVER={SQL Server};SERVER=humaxvina.cgmpdxzv57zl.ap-southeast-1.rds.amazonaws.com,1433;'
                          'DATABASE=VinaHR_Pro_Test; UID=webdev; PWD=webdev@!')
    cur = conn.cursor()
    cur.execute(query)
    return cur.fetchall()


def querydb_update(query):
    conn = pyodbc.connect('DRIVER={SQL Server};SERVER=humaxvina.cgmpdxzv57zl.ap-southeast-1.rds.amazonaws.com,1433;'
                          'DATABASE=VinaHR_Pro_Test; UID=webdev; PWD=webdev@!')
    cur = conn.cursor()
    cur.execute(query)
    cur.commit()


def get_func_name():
    return inspect.stack()[1][3]


def write_actual_excel(list_steps, func_name, duration):
    report_xlsx = report_excel
    str_steps = ''
    for i in list_steps:
        str_steps += (i + '\n')
    wb = openpyxl.load_workbook(report_xlsx)
    ws = wb.active
    # Write list steps
    for i in range(2, ws.max_row + 1):
        if func_name == ws.cell(i, 1).value:
            ws.cell(i, 10).value = str_steps
            # Write duration
            ws.cell(i, 11).value = duration
            break
    wb.save(report_xlsx)


def convert_rbg_hex(rgb):
    a = rgb.split('(')[1].split(')')[0].split(',')
    return '#{:02x}{:02x}{:02x}'.format(int(a[0]), int(a[1]), int(a[2]))


def check_radio_tick(driver, css_element):
    result = driver.find_element_by_css_selector(css_element).get_attribute('checked')
    return result


def login(driver, url):
    driver.get(url)
    time.sleep(1)
    driver.find_element_by_id('login').send_keys(user)
    driver.find_element_by_id('senha').send_keys(pass_word)
    driver.find_element_by_xpath('//button[@value="Entrar"]').click()
    time.sleep(1)


def api_login():
    password_base64 = utils.common.base64encode(pass_word)
    # login and get token
    url_login = "http://192.168.0.1/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }

    res_login = requests.post(url=url_login, json=data)
    json_data = json.loads(res_login.text)
    return json_data


def api_gateway_about():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/gateway/about"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)

    json_data_about = json.loads(res.text)
    return json_data_about


def api_network_docsis():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/network/docsis"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data_docsis = json.loads(res.text)
    return json_data_docsis


def api_network_docsis_initial_frequency():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/network/docsis/initialFrequency"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_network_docsis_procedure():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/network/docsis/procedure"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_network_docsis_channel():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/network/docsis/channel"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_network_wan():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/network/wan"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_mta_startup_procedure():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/network/docsis/mta/procedure"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_mta_line_status():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/network/docsis/mta/line"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_network_lan():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/network/lan"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_network_lan_restoreDefault():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/network/lan/ipv6/restoreDefault"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_lan_ipv6():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/gateway/devices?interface=lan-2.4g-5g"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def get_wan_info():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/network/wan"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def get_lan_info():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/network/lan"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def get_lan_v6_info():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/gateway/devices?interface=lan-2.4g-5g"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_dmz():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/dmz"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_ddns():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/ddns"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_networkOption():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/networkOption"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_upnp():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/upnp"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_firewall():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/security/firewall"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_ipv4_firewall():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/security/firewall/ipv4"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_firewall_alert():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/security/firewall/alert"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_service_log():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/log"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_service_port_ip():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/portIpFiltering"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_service_port_ip_delete_all():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/portIpFiltering/deletes"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    requests.post(url=url_request, headers=headers)


def api_service_mac_filtering():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/macFiltering"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_service_macfiltering_delete_all():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/macFiltering/deletes"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    requests.post(url=url_request, headers=headers)


def api_service_port_filtering():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/portFiltering"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_service_portfiltering_delete_all():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/portFiltering/deletes"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    requests.post(url=url_request, headers=headers)


def api_portForWarding():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/portForwarding"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_portTriggering():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/portTriggering"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_accessADM():
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/service/accessADM"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_wifi_radio(id):
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/wifi/"+str(id)+"/radio"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data


def api_wifi_scanResult(id):
    password_base64 = utils.common.base64encode(pass_word)
    url_login = url + "/api/v1/gateway/users/login"
    data = {
        "userName": user,
        "password": password_base64
    }
    res_login = requests.post(url=url_login, json=data)
    token = json.loads(res_login.text)["accessToken"]

    url_request = url + "/api/v1/wifi/0/scanResult"
    headers = {
        "content-type": "application/json",
        "content-language": "en",
        "access-token": token
    }
    res = requests.get(url=url_request, headers=headers)
    json_data = json.loads(res.text)
    return json_data
